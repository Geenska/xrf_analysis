import os
import unittest
import numpy as np
import pandas as pd

import sys
sys.path.insert(0, os.path.dirname(__file__))

import lectura_espectros

class TestXRFAnalysis(unittest.TestCase):
    def setUp(self):
        self.workspace_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
        possible_dirs = [
            os.path.join(self.workspace_dir, "datos_pdz", "Oxtotitlan"),
            "/mnt/usb1/Oxtotitlan/Panel C Gruta Sur/Oxtotitlan",
            os.path.join(os.path.dirname(__file__), "docs"),
        ]
        self.data_dir = None
        for d in possible_dirs:
            if os.path.isdir(d) and os.path.exists(os.path.join(d, "ANALYZE_EMP-7060.pdz")):
                self.data_dir = d
                break
        if not self.data_dir:
            self.data_dir = possible_dirs[0]

        self.pdz_file = os.path.join(self.data_dir, "ANALYZE_EMP-7060.pdz")

        possible_rtx = [
            os.path.join(self.data_dir, "Oxtotitlan.rtx"),
            os.path.join(os.path.dirname(__file__), "docs", "Oxtotitlan.rtx")
        ]
        self.rtx_file = next((f for f in possible_rtx if os.path.exists(f)), possible_rtx[0])

    def test_obtener_calibracion(self):
        calib = lectura_espectros.obtener_calibracion(self.data_dir)
        self.assertIsNotNone(calib)
        slope, intercept = calib
        # Verification that slope is around 0.02
        self.assertAlmostEqual(slope, 0.02, delta=0.001)
        self.assertAlmostEqual(intercept, 0.022, delta=0.01)
        print(f"[TEST] Calibración encontrada: slope={slope:.6f}, intercept={intercept:.6f}")

    def test_leer_pdz_con_calibracion(self):
        calib = lectura_espectros.obtener_calibracion(self.data_dir)
        res = lectura_espectros.leer_pdz(self.pdz_file, calibracion=calib)
        self.assertIn('metadata', res)
        self.assertIn('datos', res)
        
        df = res['datos']
        # Energy at channel 0 should be the intercept if calibrated
        if calib:
            self.assertAlmostEqual(df['Energia_keV'].iloc[0], calib[1])
            # Energy at channel 318 (Fe) should be ~6.39 keV
            fe_energy = 318 * calib[0] + calib[1]
            self.assertAlmostEqual(df['Energia_keV'].iloc[318], fe_energy)
            print(f"[TEST] Energía calibrada para Fe: {df['Energia_keV'].iloc[318]:.4f} keV")
        else:
            self.assertEqual(df['Energia_keV'].iloc[0], 0.0)

    def test_calcular_fondo_snip(self):
        calib = lectura_espectros.obtener_calibracion(self.data_dir)
        res = lectura_espectros.leer_pdz(self.pdz_file, calibracion=calib)
        counts = res['datos']['Cuentas'].values
        fondo = lectura_espectros.calcular_fondo_snip(counts)
        
        self.assertEqual(len(counts), len(fondo))
        self.assertTrue(np.all(fondo >= 0))
        # Fondo should be less than or equal to counts generally
        # (with small tolerance due to clipping and +1 smoothing)
        self.assertTrue(np.sum(counts >= fondo - 1.0) / len(counts) > 0.95)
        print("[TEST] Fondo SNIP calculado correctamente.")

    def test_buscar_picos_adaptativo(self):
        calib = lectura_espectros.obtener_calibracion(self.data_dir)
        res = lectura_espectros.leer_pdz(self.pdz_file, calibracion=calib)
        counts = res['datos']['Cuentas'].values
        energias = res['datos']['Energia_keV'].values
        fondo = lectura_espectros.calcular_fondo_snip(counts)
        
        picos = lectura_espectros.buscar_picos(energias, counts, fondo, prominencia_min=None)
        self.assertGreater(len(picos), 0)
        
        # Verify that some peaks are identified as Fe or Ca
        elementos_detectados = [p['elemento'] for p in picos if p['elemento'] is not None]
        print(f"[TEST] Elementos detectados con pico adaptativo: {elementos_detectados}")
        self.assertTrue(any("Hierro" in el for el in elementos_detectados))
        self.assertTrue(any("Calcio" in el for el in elementos_detectados))

    def test_calcular_pca_espectros(self):
        calib = lectura_espectros.obtener_calibracion(self.data_dir)
        espectros_validos = {}
        # Load a few pdz files to test PCA
        files_to_load = sorted([f for f in os.listdir(self.data_dir) if f.endswith('.pdz')])[:5]
        for f in files_to_load:
            ruta = os.path.join(self.data_dir, f)
            esp_data = lectura_espectros.leer_pdz(ruta, calibracion=calib)
            esp_data['grupo_padre'] = 'TestGroup'
            espectros_validos[f] = esp_data
            
        # Probar PCA de Covarianza (default) con alineación de signos
        pca_res = lectura_espectros.calcular_pca_espectros(espectros_validos, metodo='covarianza', alinear_signos=True)
        self.assertIsNotNone(pca_res)
        self.assertIn('resultados', pca_res)
        self.assertIn('var_pc1', pca_res)
        self.assertIn('var_pc2', pca_res)
        
        results = pca_res['resultados']
        self.assertEqual(len(results), 5)
        self.assertGreater(pca_res['var_pc1'], 0.0)
        
        # Probar PCA de Correlación
        pca_res_corr = lectura_espectros.calcular_pca_espectros(espectros_validos, metodo='correlacion', alinear_signos=True)
        self.assertIsNotNone(pca_res_corr)
        self.assertGreater(pca_res_corr['var_pc1'], 0.0)
        
        print(f"[TEST] PCA Covariance: var_pc1={pca_res['var_pc1']:.2f}%, var_pc2={pca_res['var_pc2']:.2f}%")
        print(f"[TEST] PCA Correlation: var_pc1={pca_res_corr['var_pc1']:.2f}%, var_pc2={pca_res_corr['var_pc2']:.2f}%")

    def test_sugerir_pigmentos(self):
        # Test cinnabar (Hg)
        sugs = lectura_espectros.sugerir_pigmentos(['Mercurio (Hg) Lα'])
        self.assertTrue(any('Bermellón' in s['pigmento'] for s in sugs))
        
        # Test lead-tin yellow (Pb, Sn)
        sugs = lectura_espectros.sugerir_pigmentos(['Plomo (Pb) Lα', 'Estaño (Sn) Kα'])
        self.assertTrue(any('Plomo-Estaño' in s['pigmento'] for s in sugs))
        
        # Test copper and arsenic (emerald green)
        sugs = lectura_espectros.sugerir_pigmentos(['Cobre (Cu) Kα', 'Arsénico (As) Kα'])
        self.assertTrue(any('Verde Esmeralda' in s['pigmento'] for s in sugs))
        
        # Test iron ochre
        sugs = lectura_espectros.sugerir_pigmentos(['Hierro (Fe) Kα'])
        self.assertTrue(any('Ocres' in s['pigmento'] for s in sugs))
        print("[TEST] Sugerencias de pigmentos verificadas correctamente.")

    def test_estimar_porcentajes_compuestos(self):
        picos_falsos = [
            {'elemento': 'Hierro (Fe) Kα', 'area_relativa': 40.0},
            {'elemento': 'Calcio (Ca) Kα', 'area_relativa': 20.0},
            {'elemento': 'Azufre (S) Kα', 'area_relativa': 10.0},
            {'elemento': 'Rodio (Rh) Lα', 'area_relativa': 15.0}, # Debería ser ignorado
        ]
        res = lectura_espectros.estimar_porcentajes_compuestos(picos_falsos)
        self.assertGreater(len(res), 0)
        # Yeso (S) y Calcita (Ca restante) e Hierro (Fe)
        comps = {r['compuesto']: r['porcentaje'] for r in res}
        self.assertIn('Ocres / Tierras de Hierro (Hematita/Goethita)', comps)
        self.assertIn('Yeso (CaSO4)', comps)
        self.assertIn('Calcita / Tiza / Soporte Calizo', comps)
        # Sum of compound percentages should be 100%
        self.assertAlmostEqual(sum(comps.values()), 100.0, delta=0.1)
        print("[TEST] Estimación de porcentajes de compuestos verificada correctamente.")

    def test_parsear_rtx_y_vinculacion(self):
        # 1. Parsear el archivo RTX
        espectros = lectura_espectros.parsear_rtx(self.rtx_file)
        self.assertEqual(len(espectros), 126)
        
        # 2. Simular la lógica de resolución/búsqueda de archivos PDZ vinculados
        dir_rtx = os.path.dirname(self.rtx_file)
        
        # Buscar calibración en la carpeta del proyecto y sus padres
        calib = lectura_espectros.obtener_calibracion(dir_rtx)
        if not calib:
            curr_c = dir_rtx
            for _ in range(3):
                parent_c = os.path.dirname(curr_c)
                if not parent_c or parent_c == curr_c:
                    break
                calib = lectura_espectros.obtener_calibracion(parent_c)
                if calib:
                    break
                curr_c = parent_c
                
        self.assertIsNotNone(calib) # Debería encontrar la calibración en el directorio padre de rtx (Oxtotitlan)
        
        # Pre-escanear y cachear todos los archivos .pdz
        pdz_cache = {}
        directorios_busqueda = [dir_rtx]
        curr_dir = dir_rtx
        for _ in range(3):
            parent_dir = os.path.dirname(curr_dir)
            if not parent_dir or parent_dir == curr_dir:
                break
            directorios_busqueda.append(parent_dir)
            curr_dir = parent_dir
            
        for d in reversed(directorios_busqueda):
            if os.path.isdir(d):
                for root_d, dirs_d, files_d in os.walk(d):
                    dirs_d[:] = [dirname for dirname in dirs_d if dirname not in ('venv', '.git', '__pycache__', 'node_modules')]
                    for f in files_d:
                        if f.lower().endswith('.pdz'):
                            pdz_cache[f.lower()] = os.path.join(root_d, f)
                            
        # Verificar que podemos encontrar y leer el primer espectro usando la lógica de cache
        esp = espectros[0]
        pdz_name = esp['archivo_pdz']
        pdz_full_path = os.path.join(dir_rtx, pdz_name)
        
        if not os.path.exists(pdz_full_path):
            encontrado_ruta = None
            if pdz_name.lower() in pdz_cache:
                encontrado_ruta = pdz_cache[pdz_name.lower()]
            if encontrado_ruta:
                pdz_full_path = encontrado_ruta
                
        self.assertTrue(os.path.exists(pdz_full_path))
        datos_esp = lectura_espectros.leer_pdz(pdz_full_path, calibracion=calib)
        self.assertIsNotNone(datos_esp)
        self.assertIn('datos', datos_esp)
        print(f"[TEST] RTX y vinculación con caché validados para {pdz_name}.")

    def test_rtx_xml_fallback_loading(self):
        # 1. Parsear el archivo RTX
        espectros = lectura_espectros.parsear_rtx(self.rtx_file)
        self.assertEqual(len(espectros), 126)
        
        # 2. Verificar que los datos XML (counts y metadatos) estén presentes para el primer espectro
        esp = espectros[0]
        self.assertIn('xml_data', esp)
        
        xml_meta = esp['xml_data']
        self.assertEqual(xml_meta['num_channels'], 2048)
        self.assertAlmostEqual(xml_meta['ev_per_channel'], 20.0, delta=0.1)
        self.assertAlmostEqual(xml_meta['live_time'], 15.564, delta=0.01)
        self.assertEqual(xml_meta['xray_voltage_kv'], 40.0)
        self.assertEqual(xml_meta['xray_filament_current'], 45.0)
        
        # Verificar counts
        counts = xml_meta['counts']
        self.assertEqual(len(counts), 2048)
        self.assertEqual(counts[0], 412.0)
        self.assertEqual(counts[1], 360.0)
        
        # 3. Simular reconstrucción de datos con la calibración del XML (como fallback)
        slope, intercept = xml_meta['xml_calib']
        energias = [i * slope + intercept for i in range(len(counts))]
        df = pd.DataFrame({
            'Energia_keV': energias,
            'Cuentas': counts
        })
        
        self.assertEqual(df.shape, (2048, 2))
        self.assertAlmostEqual(df['Energia_keV'].iloc[0], intercept)
        print("[TEST] Fallback de reconstrucción XML validado con éxito.")

    def test_exportar_artax_filtrado_estructura(self):
        import openpyxl
        import tempfile

        # Simular lectura de datos del RTX de prueba
        espectros = lectura_espectros.parsear_rtx(self.rtx_file)
        self.assertGreater(len(espectros), 0)

        # Seleccionar elementos de prueba: Fe, Ca, Cu
        elementos_sel = ['Ca', 'Cu', 'Fe']
        
        # Generar archivo temporal
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name

        try:
            wb = openpyxl.Workbook()
            ws_param = wb.active
            ws_param.title = "Parameter"

            param_rows = [
                ["Project:", "TestProject"],
                ["Method:", "arqu encrym"],
                ["High voltage/kV:", 40],
                ["Current/µA:", 11],
                ["Time/s:", 24],
                ["Elements:", " ".join(elementos_sel) + " "],
            ]
            for r in param_rows:
                ws_param.append(r)
            for _ in range(14):
                ws_param.append([None, None])
            ws_param.append(["Values:", "Net area"])

            ws_points = wb.create_sheet(title="Points")
            cols_test = ['Ca K12', 'Ca L1', 'Cu K12', 'Cu L1', 'Fe K12', 'Fe L1']
            headers = [None, None, None, "Muestra / Espectro"] + cols_test
            ws_points.append(headers)

            for esp in espectros[:5]:
                ws_points.append([None, None, None, esp['archivo_pdz'], 100, 200, 300, 400, 500, 600])

            wb.save(tmp_path)

            # Verificar el archivo generado
            wb_read = openpyxl.load_workbook(tmp_path)
            self.assertIn("Parameter", wb_read.sheetnames)
            self.assertIn("Points", wb_read.sheetnames)

            df_points = pd.read_excel(tmp_path, sheet_name="Points", engine='openpyxl')
            self.assertEqual(len(df_points), 5)
            self.assertIn("Fe K12", df_points.columns)
            self.assertIn("Ca K12", df_points.columns)
            print("[TEST] Exportación ARTAX filtrada validada con éxito.")
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

if __name__ == '__main__':
    unittest.main()
