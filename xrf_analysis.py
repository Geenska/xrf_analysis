import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
from tkinterdnd2 import DND_FILES, TkinterDnD

import lectura_espectros
import openpyxl

# Diccionario de asociaciones de elementos comunes en patrimonio cultural, aleaciones y pigmentos
ASOCIACIONES_PATRIMONIO = {
    # Metales nobles y moneda/orfebrería
    'Au': ['Ag', 'Cu', 'Pt', 'Hg', 'Pb'],
    'Ag': ['Au', 'Cu', 'Pb', 'Cl', 'S', 'Sb'],
    'Pt': ['Au', 'Ag', 'Fe'],
    
    # Bronces, latones y aleaciones metálicas arqueológicas
    'Cu': ['Sn', 'Zn', 'Pb', 'As', 'Ag', 'Au', 'Ni', 'Sb', 'Bi', 'Cl', 'S', 'Si', 'Ca'],
    'Sn': ['Cu', 'Pb', 'Sb', 'Si', 'Fe'],
    'Zn': ['Cu', 'Ba', 'Pb', 'Ti'],
    'Pb': ['Sn', 'Sb', 'As', 'Cu', 'Ag', 'S', 'Ca', 'Bi'],
    'Sb': ['Pb', 'Cu', 'Sn', 'Ca'],
    'As': ['Cu', 'S', 'Pb'],
    'Bi': ['Pb', 'Cu', 'Sn', 'Co'],
    'Ni': ['Cu', 'Fe', 'Co'],
    
    # Hierro, tierras, pigmentos de tierra y corrosión
    'Fe': ['Mn', 'Si', 'Al', 'K', 'Ca', 'Ti', 'S', 'P', 'Cl'],
    'Mn': ['Fe', 'Ba', 'Ca', 'Si'],
    'Ti': ['Fe', 'Zn', 'Ba', 'Si', 'Ca'],
    'Al': ['Si', 'K', 'Fe', 'Ca'],
    'Si': ['Al', 'K', 'Ca', 'Fe', 'Pb'],
    'K':  ['Si', 'Al', 'Fe', 'Ca'],
    
    # Preparaciones, cargas, aglutinantes y pigmentos minerales
    'Ca': ['S', 'P', 'Sr', 'Si', 'Al', 'Pb', 'Mg'],
    'S':  ['Ca', 'Hg', 'As', 'Cu', 'Pb', 'Ba', 'Fe'],
    'P':  ['Ca', 'Fe'],
    'Sr': ['Ca', 'Ba', 'S'],
    'Ba': ['S', 'Zn', 'Mn', 'Ti'],
    
    # Pigmentos específicos y elementos de conservación
    'Hg': ['S', 'Au', 'Pb'],
    'Co': ['Ni', 'As', 'Fe', 'Bi', 'Si', 'K'],
    'Cr': ['Pb', 'Fe', 'Zn'],
    'Cd': ['S', 'Zn'],
    'Cl': ['Cu', 'Pb', 'Ag'],
    'Br': ['Ag', 'Pb'],
    'Zr': ['Si', 'Ti'],
    'Mg': ['Ca', 'Si', 'Al']
}

def obtener_elementos_relevantes_patrimonio(elementos_detectados):
    """
    Retorna (detectados, asociados, todos_relevantes) a partir de los elementos detectados
    utilizando el mapa de asociaciones de patrimonio cultural.
    """
    detectados = set(elementos_detectados)
    asociados = set()
    for el in detectados:
        if el in ASOCIACIONES_PATRIMONIO:
            for assoc in ASOCIACIONES_PATRIMONIO[el]:
                if assoc not in detectados:
                    asociados.add(assoc)
                    
    elementos_validos = set(key.split('_')[0] for key in lectura_espectros.XRF_ELEMENTS.keys())
    detectados = detectados.intersection(elementos_validos)
    asociados = asociados.intersection(elementos_validos)
    
    return detectados, asociados, detectados.union(asociados)

class XRFProcessorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Laboratorio CODICE - Espectroscopia XRF")
        self.root.geometry("1280x750")
        
        self.current_theme = "light"
        self.definir_colores()
        
        self.root.configure(bg=self.BG_MAIN)
        
        # Almacenamiento de datos de espectros cargados
        # Llave: ID único en el Treeview, Valor: dict con {'datos': DataFrame, 'metadata': dict}
        self.espectros_datos = {}
        self.mostrar_fondo_y_picos = False
        
        # Configurar estilos visuales limpios (clam)
        self.style = ttk.Style()
        self.style.theme_use('clam')
        
        self.aplicar_estilos()
        
        self.crear_componentes()
        self.configurar_eventos()

    def definir_colores(self):
        if self.current_theme == "dark":
            # Paleta de colores Premium Dark Theme
            self.BG_MAIN = "#121214"
            self.BG_PANEL = "#1c1c21"
            self.BG_CARD = "#25252b"
            self.FG_MAIN = "#f3f4f6"
            self.FG_MUTED = "#9ca3af"
            self.ACCENT = "#4f46e5"       # Indigo-600
            self.ACCENT_HOVER = "#6366f1" # Indigo-500
            self.SUCCESS = "#059669"      # Emerald-600
            self.BORDER = "#2e2e38"
        else:
            # Paleta de colores Premium Light Theme
            self.BG_MAIN = "#f3f4f6"      # Gris claro
            self.BG_PANEL = "#ffffff"     # Blanco
            self.BG_CARD = "#f9fafb"      # Blanco cálido
            self.FG_MAIN = "#111827"      # Gris oscuro / negro
            self.FG_MUTED = "#4b5563"     # Gris atenuado
            self.ACCENT = "#4f46e5"       # Indigo-600 idéntico
            self.ACCENT_HOVER = "#3730a3" # Indigo más oscuro para hover
            self.SUCCESS = "#059669"      # Verde éxito
            self.BORDER = "#d1d5db"       # Borde gris claro

    def aplicar_estilos(self):
        # Aplicar colores al estilo ttk
        self.style.configure('.', background=self.BG_PANEL, foreground=self.FG_MAIN, fieldbackground=self.BG_PANEL, bordercolor=self.BORDER, font=('Helvetica', 10))
        self.style.configure('TFrame', background=self.BG_PANEL)
        self.style.configure('TLabel', background=self.BG_PANEL, foreground=self.FG_MAIN)
        
        # Checkbuttons
        self.style.configure('TCheckbutton', background=self.BG_PANEL, foreground=self.FG_MAIN)
        self.style.map('TCheckbutton',
            background=[('active', self.BG_PANEL)],
            foreground=[('active', self.ACCENT_HOVER)]
        )
        
        # Botones personalizados
        self.style.configure('Action.TButton', background=self.ACCENT, foreground=self.FG_MAIN, bordercolor=self.BORDER, font=('Helvetica', 10, 'bold'))
        self.style.map('Action.TButton',
            background=[('active', self.ACCENT_HOVER), ('pressed', self.ACCENT)],
            foreground=[('active', self.FG_MAIN)]
        )
        
        self.style.configure('Success.TButton', background=self.SUCCESS, foreground=self.FG_MAIN, bordercolor=self.BORDER, font=('Helvetica', 10, 'bold'))
        self.style.map('Success.TButton',
            background=[('active', '#059669'), ('pressed', self.SUCCESS)],
            foreground=[('active', self.FG_MAIN)]
        )
        
        self.style.configure('Danger.TButton', background='#b91c1c', foreground=self.FG_MAIN, bordercolor=self.BORDER, font=('Helvetica', 10, 'bold'))
        self.style.map('Danger.TButton',
            background=[('active', '#dc2626'), ('pressed', '#b91c1c')],
            foreground=[('active', self.FG_MAIN)]
        )
        
        # Botón especial para alternar tema
        self.style.configure('Theme.TButton', background=self.BG_CARD, foreground=self.FG_MAIN, bordercolor=self.BORDER, font=('Helvetica', 10, 'bold'))
        self.style.map('Theme.TButton',
            background=[('active', self.BG_PANEL)],
            foreground=[('active', self.ACCENT_HOVER)]
        )
        
        # Labelframe
        self.style.configure('TLabelframe', background=self.BG_PANEL, foreground=self.FG_MAIN, bordercolor=self.BORDER)
        self.style.configure('TLabelframe.Label', background=self.BG_PANEL, foreground=self.ACCENT_HOVER, font=('Helvetica', 10, 'bold'))
        
        # Notebook (Pestañas)
        self.style.configure('TNotebook', background=self.BG_MAIN, bordercolor=self.BORDER, tabmargins=[2, 5, 2, 0])
        self.style.configure('TNotebook.Tab', background=self.BG_CARD, foreground=self.FG_MUTED, bordercolor=self.BORDER, font=('Helvetica', 10, 'bold'), padding=(15, 6))
        self.style.map('TNotebook.Tab',
            background=[('selected', self.BG_PANEL)],
            foreground=[('selected', self.ACCENT_HOVER)]
        )
        
        # Treeview (Listado de archivos)
        self.style.configure('Treeview', background=self.BG_CARD, foreground=self.FG_MAIN, fieldbackground=self.BG_CARD, bordercolor=self.BORDER, rowheight=24)
        self.style.configure('Treeview.Heading', background=self.BG_PANEL, foreground=self.FG_MAIN, bordercolor=self.BORDER, font=('Helvetica', 9, 'bold'))
        self.style.map('Treeview',
            background=[('selected', self.ACCENT)],
            foreground=[('selected', self.FG_MAIN)]
        )
        
        # Combobox
        self.style.configure('TCombobox', background=self.BG_CARD, foreground=self.FG_MAIN, fieldbackground=self.BG_CARD, bordercolor=self.BORDER)
        self.style.map('TCombobox',
            fieldbackground=[('readonly', self.BG_CARD)],
            selectbackground=[('readonly', self.ACCENT)]
        )

    def toggle_theme(self):
        self.current_theme = "light" if self.current_theme == "dark" else "dark"
        self.definir_colores()
        self.aplicar_estilos()
        
        # Actualizar widgets manuales
        self.root.configure(bg=self.BG_MAIN)
        if hasattr(self, 'left_canvas'):
            self.left_canvas.configure(bg=self.BG_MAIN)
        
        # Botón de tema
        self.btn_theme.config(text="☀️ Tema Claro" if self.current_theme == "dark" else "🌙 Tema Oscuro")
        
        # Drop zone label
        self.drop_label.configure(
            bg=self.BG_CARD, 
            fg=self.FG_MUTED,
            highlightbackground=self.BORDER
        )
        
        # Separador cabecera
        self.header_sep.configure(bg=self.BORDER)
        
        # Text widgets
        self.pigment_text.configure(
            bg=self.BG_CARD,
            fg=self.FG_MAIN,
            insertbackground=self.FG_MAIN,
            highlightbackground=self.BORDER
        )
        self.guia_text.configure(
            bg=self.BG_CARD,
            fg=self.FG_MAIN,
            insertbackground=self.FG_MAIN,
            highlightbackground=self.BORDER
        )
        
        # Matplotlib figures
        self.fig.set_facecolor(self.BG_PANEL)
        self.ax.set_facecolor(self.BG_CARD)
        self.ax.xaxis.label.set_color(self.FG_MUTED)
        self.ax.yaxis.label.set_color(self.FG_MUTED)
        self.ax.title.set_color(self.FG_MAIN)
        self.ax.tick_params(colors=self.FG_MUTED, which='both')
        for spine in self.ax.spines.values():
            spine.set_color(self.BORDER)
        self.ax.grid(True, linestyle="--", alpha=0.15, color=self.FG_MUTED)
        
        self.fig_pca.set_facecolor(self.BG_PANEL)
        self.ax_pca.set_facecolor(self.BG_CARD)
        self.ax_pca.xaxis.label.set_color(self.FG_MUTED)
        self.ax_pca.yaxis.label.set_color(self.FG_MUTED)
        self.ax_pca.title.set_color(self.FG_MAIN)
        self.ax_pca.tick_params(colors=self.FG_MUTED, which='both')
        for spine in self.ax_pca.spines.values():
            spine.set_color(self.BORDER)
        self.ax_pca.grid(True, linestyle="--", alpha=0.15, color=self.FG_MUTED)
        
        # Toolbar matplotlib config
        try:
            self.toolbar.config(background=self.BG_PANEL)
            for button in self.toolbar.winfo_children():
                try:
                    button.config(background=self.BG_PANEL, foreground=self.FG_MAIN)
                except Exception:
                    pass
        except Exception:
            pass
            
        # Menu contextual config
        try:
            self.menu_contextual.configure(
                bg=self.BG_CARD,
                fg=self.FG_MAIN,
                activebackground=self.ACCENT,
                activeforeground=self.FG_MAIN
            )
        except Exception:
            pass
            
        # Re-plot or draw to apply colors immediately
        try:
            self.discern_canvas.configure(bg=self.BG_PANEL)
        except Exception:
            pass
        self.canvas.draw()
        self.canvas_pca.draw()

    def crear_componentes(self):
        # Configurar la ventana principal (self.root) con una estructura de filas
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=0)  # Cabecera
        self.root.rowconfigure(1, weight=0)  # Separador
        self.root.rowconfigure(2, weight=1)  # Contenedor principal

        # -------------------------------------------------------------
        # CABECERA SUPERIOR (HEADER PANEL)
        # -------------------------------------------------------------
        self.header_frame = ttk.Frame(self.root, padding=(20, 12))
        self.header_frame.grid(row=0, column=0, sticky="ew")
        self.header_frame.columnconfigure(0, weight=1)
        self.header_frame.columnconfigure(1, weight=0)

        self.header_label = ttk.Label(
            self.header_frame, 
            text="Laboratorio CODICE - Espectroscopia XRF", 
            font=("Helvetica", 16, "bold"), 
            foreground=self.ACCENT_HOVER
        )
        self.header_label.grid(row=0, column=0, sticky="w")

        self.subtitle_label = ttk.Label(
            self.header_frame, 
            text="CNCPC - Procesamiento Masivo e Interpretación de Pigmentos Históricos y Compuestos Arqueológicos", 
            font=("Helvetica", 9, "italic"), 
            foreground=self.FG_MUTED
        )
        self.subtitle_label.grid(row=1, column=0, sticky="w", pady=(2, 0))

        # Botón para alternar tema (Oscuro / Claro)
        self.btn_theme = ttk.Button(
            self.header_frame,
            text="☀️ Tema Claro" if self.current_theme == "dark" else "🌙 Tema Oscuro",
            command=self.toggle_theme,
            style='Theme.TButton'
        )
        self.btn_theme.grid(row=0, column=1, rowspan=2, sticky="e", padx=5)

        # Línea divisoria
        self.header_sep = tk.Frame(self.root, height=1, bg=self.BORDER, bd=0, highlightthickness=0)
        self.header_sep.grid(row=1, column=0, sticky="ew")

        # Contenedor Principal (Main container)
        self.main_container = ttk.Frame(self.root)
        self.main_container.grid(row=2, column=0, sticky="nsew")
        self.main_container.columnconfigure(0, weight=1)  # Panel izquierdo (1/4)
        self.main_container.columnconfigure(1, weight=3)  # Panel derecho (3/4)
        self.main_container.rowconfigure(0, weight=1)

        # -------------------------------------------------------------
        # PANEL IZQUIERDO: CONTROLES, ÁRBOL Y REFERENCIAS (DESPLAZABLE)
        # -------------------------------------------------------------
        self.left_container = ttk.Frame(self.main_container)
        self.left_container.grid(row=0, column=0, sticky="nsew", padx=(10, 5), pady=10)
        self.left_container.rowconfigure(0, weight=1)
        self.left_container.columnconfigure(0, weight=1)

        self.left_canvas = tk.Canvas(
            self.left_container, 
            bg=self.BG_MAIN, 
            bd=0, 
            highlightthickness=0
        )
        self.left_scrollbar = ttk.Scrollbar(
            self.left_container, 
            orient="vertical", 
            command=self.left_canvas.yview
        )

        self.left_frame = ttk.Frame(self.left_canvas, padding=5)
        self.left_frame.columnconfigure(0, weight=1)

        self.left_frame_window = self.left_canvas.create_window(
            (0, 0), 
            window=self.left_frame, 
            anchor="nw"
        )

        def _on_canvas_configure(event):
            self.left_canvas.itemconfig(self.left_frame_window, width=event.width)

        def _on_frame_configure(event):
            self.left_canvas.configure(scrollregion=self.left_canvas.bbox("all"))

        self.left_canvas.bind("<Configure>", _on_canvas_configure)
        self.left_frame.bind("<Configure>", _on_frame_configure)
        self.left_canvas.configure(yscrollcommand=self.left_scrollbar.set)

        self.left_canvas.grid(row=0, column=0, sticky="nsew")
        self.left_scrollbar.grid(row=0, column=1, sticky="ns")

        # Vincular rueda del ratón al canvas al pasar el cursor encima
        def _on_mousewheel(event):
            if event.delta:
                self.left_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            elif event.num == 4:
                self.left_canvas.yview_scroll(-1, "units")
            elif event.num == 5:
                self.left_canvas.yview_scroll(1, "units")

        def _bind_mousewheel(event):
            self.left_canvas.bind_all("<MouseWheel>", _on_mousewheel)
            self.left_canvas.bind_all("<Button-4>", _on_mousewheel)
            self.left_canvas.bind_all("<Button-5>", _on_mousewheel)

        def _unbind_mousewheel(event):
            self.left_canvas.unbind_all("<MouseWheel>")
            self.left_canvas.unbind_all("<Button-4>")
            self.left_canvas.unbind_all("<Button-5>")

        self.left_canvas.bind("<Enter>", _bind_mousewheel)
        self.left_canvas.bind("<Leave>", _unbind_mousewheel)

        # 1. Zona de Arrastre (Drop Zone)
        self.drop_label = tk.Label(
            self.left_frame, 
            text="Arrastra y suelta aquí tus archivos\n.pdz o tu proyecto .rtx de Artax",
            bg=self.BG_CARD, 
            fg=self.FG_MUTED,
            relief="flat", 
            bd=0,
            font=("Helvetica", 11, "italic"),
            height=3,
            highlightthickness=1,
            highlightbackground=self.BORDER,
            highlightcolor=self.ACCENT
        )
        self.drop_label.grid(row=0, column=0, sticky="ew", pady=(0, 5))
        
        # Registrar eventos de arrastre
        self.drop_label.drop_target_register(DND_FILES)
        self.drop_label.dnd_bind('<<Drop>>', self.procesar_drop)

        # 1.5 Botones para buscar manualmente archivos y carpetas
        self.load_buttons_frame = ttk.Frame(self.left_frame)
        self.load_buttons_frame.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        self.load_buttons_frame.columnconfigure(0, weight=1)
        self.load_buttons_frame.columnconfigure(1, weight=1)

        self.btn_buscar = ttk.Button(
            self.load_buttons_frame,
            text="📂 Buscar Archivos...",
            command=self.seleccionar_archivos,
            style='Action.TButton'
        )
        self.btn_buscar.grid(row=0, column=0, sticky="ew", padx=(0, 2))

        self.btn_buscar_carpeta = ttk.Button(
            self.load_buttons_frame,
            text="📁 Cargar Carpeta...",
            command=self.seleccionar_carpeta,
            style='Action.TButton'
        )
        self.btn_buscar_carpeta.grid(row=0, column=1, sticky="ew", padx=(2, 0))

        # 1.6 Buscador para filtrado en el árbol
        self.search_frame = ttk.Frame(self.left_frame)
        self.search_frame.grid(row=2, column=0, sticky="ew", pady=(0, 5))
        self.search_frame.columnconfigure(1, weight=1)
        
        lbl_search = ttk.Label(self.search_frame, text="🔍 Buscar:", font=("Helvetica", 9))
        lbl_search.grid(row=0, column=0, padx=(0, 5))
        
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *args: self.on_search_change())
        self.search_entry = tk.Entry(
            self.search_frame,
            textvariable=self.search_var,
            font=("Helvetica", 9),
            bg=self.BG_CARD,
            fg=self.FG_MAIN,
            insertbackground=self.FG_MAIN,
            relief="flat",
            highlightthickness=1,
            highlightbackground=self.BORDER,
            highlightcolor=self.ACCENT
        )
        self.search_entry.grid(row=0, column=1, sticky="ew")

        # 2. Contenedor de Treeview con scrollbar (Árbol de archivos)
        self.tree_frame = ttk.Frame(self.left_frame)
        self.tree_frame.grid(row=3, column=0, sticky="ew", pady=(0, 5))
        self.tree_frame.rowconfigure(0, weight=1)
        self.tree_frame.columnconfigure(0, weight=1)

        self.file_tree = ttk.Treeview(self.tree_frame, show="tree", selectmode="extended", height=8)
        self.file_tree.grid(row=0, column=0, sticky="nsew")

        self.scrollbar = ttk.Scrollbar(self.tree_frame, orient="vertical", command=self.file_tree.yview)
        self.scrollbar.grid(row=0, column=1, sticky="ns")
        self.file_tree.configure(yscrollcommand=self.scrollbar.set)

        # 2.4 Botones para control de carpetas (Crear y Mover)
        self.tree_ctrl_frame = ttk.Frame(self.left_frame)
        self.tree_ctrl_frame.grid(row=4, column=0, sticky="ew", pady=(0, 10))
        self.tree_ctrl_frame.columnconfigure(0, weight=1)
        self.tree_ctrl_frame.columnconfigure(1, weight=1)

        self.btn_crear_carpeta = ttk.Button(
            self.tree_ctrl_frame,
            text="📁+ Nueva Carpeta...",
            command=self.crear_nueva_carpeta,
            style='Theme.TButton'
        )
        self.btn_crear_carpeta.grid(row=0, column=0, sticky="ew", padx=(0, 2))

        self.btn_mover_a = ttk.Button(
            self.tree_ctrl_frame,
            text="➡️ Mover a...",
            command=self.mover_espectro_a,
            style='Theme.TButton'
        )
        self.btn_mover_a.grid(row=0, column=1, sticky="ew", padx=(2, 0))

        # Crear menú contextual para clic derecho en el Treeview
        self.menu_contextual = tk.Menu(self.file_tree, tearoff=0)
        self.menu_contextual.add_command(label="📁+ Nueva Carpeta...", command=self.crear_nueva_carpeta)
        self.menu_contextual.add_command(label="➡️ Mover a...", command=self.mover_espectro_a)
        self.menu_contextual.add_command(label="❌ Eliminar", command=self.eliminar_elemento_arbol)
        
        self.menu_contextual.configure(
            bg=self.BG_CARD,
            fg=self.FG_MAIN,
            activebackground=self.ACCENT,
            activeforeground=self.FG_MAIN,
            bd=1,
            relief="flat"
        )

        # 2.5 Panel de Líneas de Referencia de Elementos
        self.ref_frame = ttk.LabelFrame(self.left_frame, text="Líneas de Referencia (Elementos)", padding=8)
        self.ref_frame.grid(row=5, column=0, sticky="ew", pady=(0, 10))
        
        self.elementos_var = {}
        self.actualizar_panel_referencia(None)

        # 2.6 Panel de Rango de Visualización
        self.plot_ctrl_frame = ttk.LabelFrame(self.left_frame, text="Configuración del Gráfico", padding=8)
        self.plot_ctrl_frame.grid(row=6, column=0, sticky="ew", pady=(0, 10))
        self.plot_ctrl_frame.columnconfigure(3, weight=1)
        
        ttk.Label(self.plot_ctrl_frame, text="Rango Max (keV):").grid(row=0, column=0, sticky="w", padx=(5, 2))
        self.rango_max_var = tk.StringVar(value="16")
        self.combo_rango = ttk.Combobox(
            self.plot_ctrl_frame, 
            textvariable=self.rango_max_var, 
            values=["16", "30", "40", "Manual"], 
            width=7, 
            state="readonly"
        )
        self.combo_rango.grid(row=0, column=1, sticky="w", padx=(2, 2))
        self.combo_rango.bind("<<ComboboxSelected>>", self.on_rango_changed)

        # Entrada para ajuste manual de energía en keV
        self.rango_manual_var = tk.StringVar(value="20")
        self.entry_rango_manual = tk.Entry(
            self.plot_ctrl_frame,
            textvariable=self.rango_manual_var,
            width=5,
            font=("Helvetica", 9),
            bg=self.BG_CARD,
            fg=self.FG_MAIN,
            insertbackground=self.FG_MAIN,
            relief="flat",
            highlightthickness=1,
            highlightbackground=self.BORDER,
            highlightcolor=self.ACCENT
        )
        self.entry_rango_manual.bind("<Return>", lambda e: self.on_element_toggle())
        self.entry_rango_manual.bind("<FocusOut>", lambda e: self.on_element_toggle())
        self.entry_rango_manual.grid_remove()

        self.mostrar_etiquetas_picos_var = tk.BooleanVar(value=False)
        self.chk_etiquetas_picos = ttk.Checkbutton(
            self.plot_ctrl_frame,
            text="🏷️ Mostrar etiquetas de elementos",
            variable=self.mostrar_etiquetas_picos_var,
            command=self.on_element_toggle
        )
        self.chk_etiquetas_picos.grid(row=1, column=0, columnspan=3, sticky="w", padx=5, pady=(4, 0))

        # 3. Contenedor de Botones de Acción
        self.buttons_frame = ttk.Frame(self.left_frame)
        self.buttons_frame.grid(row=7, column=0, sticky="ew")
        self.buttons_frame.columnconfigure(0, weight=1)

        self.btn_procesar = ttk.Button(
            self.buttons_frame, 
            text="⚙️ Calcular Picos y Fondo", 
            command=self.toggle_procesamiento,
            style='Action.TButton'
        )
        self.btn_procesar.grid(row=0, column=0, sticky="ew", pady=4)
        
        self.btn_excel = ttk.Button(
            self.buttons_frame, 
            text="📊 Exportar a Excel Universal", 
            command=self.exportar_datos,
            style='Success.TButton'
        )
        self.btn_excel.grid(row=1, column=0, sticky="ew", pady=4)

        self.btn_excel_graficas = ttk.Button(
            self.buttons_frame, 
            text="📈 Exportar con Gráficas por Hoja", 
            command=self.exportar_datos_con_graficas,
            style='Success.TButton'
        )
        self.btn_excel_graficas.grid(row=2, column=0, sticky="ew", pady=4)

        self.btn_guardar_rtx = ttk.Button(
            self.buttons_frame, 
            text="💾 Guardar Proyecto ARTAX (.rtx)", 
            command=self.exportar_proyecto_rtx,
            style='Action.TButton'
        )
        self.btn_guardar_rtx.grid(row=3, column=0, sticky="ew", pady=4)

        self.btn_artax_filtrado = ttk.Button(
            self.buttons_frame,
            text="🎯 Exportar ARTAX (Líneas Seleccionadas)",
            command=self.exportar_artax_filtrado,
            style='Success.TButton'
        )
        self.btn_artax_filtrado.grid(row=4, column=0, sticky="ew", pady=4)

        self.btn_limpiar = ttk.Button(
            self.buttons_frame, 
            text="🧹 Limpiar Todo", 
            command=self.limpiar_datos,
            style='Danger.TButton'
        )
        self.btn_limpiar.grid(row=5, column=0, sticky="ew", pady=4)

        # -------------------------------------------------------------
        # PANEL DERECHO: PESTAÑAS (NOTEBOOK)
        # -------------------------------------------------------------
        self.notebook = ttk.Notebook(self.main_container)
        self.notebook.grid(row=0, column=1, sticky="nsew", padx=(5, 10), pady=10)

        # Pestaña 1: Espectro y Gráfica
        self.tab_espectro = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_espectro, text="Visualización del Espectro")
        self.tab_espectro.rowconfigure(0, weight=1)
        self.tab_espectro.columnconfigure(0, weight=1)

        # Crear figura de espectro
        self.fig = Figure(figsize=(7, 5), dpi=100, facecolor=self.BG_PANEL)
        self.ax = self.fig.add_subplot(111)
        self.ax.set_facecolor(self.BG_CARD)
        self.ax.set_title("Espectro XRF", color=self.FG_MAIN, weight='bold')
        self.ax.set_xlabel("Energía (keV)", color=self.FG_MUTED)
        self.ax.set_ylabel("Cuentas", color=self.FG_MUTED)
        self.ax.tick_params(colors=self.FG_MUTED, which='both')
        for spine in self.ax.spines.values():
            spine.set_color(self.BORDER)
        self.ax.grid(True, linestyle="--", alpha=0.15, color=self.FG_MUTED)

        self.canvas = FigureCanvasTkAgg(self.fig, master=self.tab_espectro)
        self.canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew")

        self.toolbar_frame = ttk.Frame(self.tab_espectro)
        self.toolbar_frame.grid(row=1, column=0, sticky="ew")
        self.toolbar = NavigationToolbar2Tk(self.canvas, self.toolbar_frame)
        self.toolbar.update()

        self.btn_copiar_portapapeles = ttk.Button(
            self.toolbar_frame,
            text="📋 Copiar Imagen al Portapapeles",
            command=self.copiar_grafica_portapapeles,
            style='Theme.TButton'
        )
        self.btn_copiar_portapapeles.pack(side=tk.LEFT, padx=10, pady=2)

        try:
            self.toolbar.config(background=self.BG_PANEL)
            for button in self.toolbar.winfo_children():
                try:
                    button.config(background=self.BG_PANEL, foreground=self.FG_MAIN)
                except Exception:
                    pass
        except Exception:
            pass

        # Pestaña 2: Tabla de Elementos Detectados (Análisis Cuantitativo)
        self.tab_cuantitativo = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_cuantitativo, text="Análisis Cuantitativo")
        self.tab_cuantitativo.rowconfigure(1, weight=3)
        self.tab_cuantitativo.rowconfigure(2, weight=1)
        self.tab_cuantitativo.columnconfigure(0, weight=1)
        
        lbl_info_tabla = ttk.Label(
            self.tab_cuantitativo, 
            text="Picos identificados en el espectro seleccionado (Sustrayendo fondo SNIP):",
            font=("Helvetica", 10, "bold"),
            padding=10
        )
        lbl_info_tabla.grid(row=0, column=0, sticky="w")

        # Configurar tabla de picos
        self.peaks_table = ttk.Treeview(
            self.tab_cuantitativo,
            columns=('elemento', 'energia', 'cuentas_brutas', 'cuentas_netas', 'area_neta', 'area_relativa'),
            show='headings'
        )
        self.peaks_table.heading('elemento', text='Elemento Identificado')
        self.peaks_table.heading('energia', text='Energía del Pico (keV)')
        self.peaks_table.heading('cuentas_brutas', text='Cuentas Brutas')
        self.peaks_table.heading('cuentas_netas', text='Cuentas Netas')
        self.peaks_table.heading('area_neta', text='Área Neta Integrada')
        self.peaks_table.heading('area_relativa', text='Intensidad Relativa %')

        self.peaks_table.column('elemento', width=180, anchor='center')
        self.peaks_table.column('energia', width=140, anchor='center')
        self.peaks_table.column('cuentas_brutas', width=140, anchor='center')
        self.peaks_table.column('cuentas_netas', width=140, anchor='center')
        self.peaks_table.column('area_neta', width=150, anchor='center')
        self.peaks_table.column('area_relativa', width=140, anchor='center')

        self.peaks_table.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 10))
        
        # Scrollbar para la tabla
        self.table_scroll = ttk.Scrollbar(self.tab_cuantitativo, orient="vertical", command=self.peaks_table.yview)
        self.table_scroll.grid(row=1, column=1, sticky="ns", pady=(0, 10))
        self.peaks_table.configure(yscrollcommand=self.table_scroll.set)

        # Panel de Interpretación de Pigmentos Sugeridos (Con Discernimiento Humano)
        self.pigment_frame = ttk.LabelFrame(self.tab_cuantitativo, text="Interpretación de Pigmentos Sugeridos (Teoría Histórica y Discernimiento Humano)", padding=10)
        self.pigment_frame.grid(row=2, column=0, columnspan=2, sticky="nsew", padx=10, pady=(0, 10))
        self.pigment_frame.rowconfigure(0, weight=1)
        self.pigment_frame.columnconfigure(0, weight=3)
        self.pigment_frame.columnconfigure(1, weight=2)
        
        # Sub-frame para texto y su scrollbar
        self.text_scroll_frame = ttk.Frame(self.pigment_frame)
        self.text_scroll_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        self.text_scroll_frame.rowconfigure(0, weight=1)
        self.text_scroll_frame.columnconfigure(0, weight=1)
        
        self.pigment_text = tk.Text(
            self.text_scroll_frame, 
            height=6, 
            font=("Helvetica", 9), 
            wrap="word", 
            bg=self.BG_CARD, 
            fg=self.FG_MAIN, 
            insertbackground=self.FG_MAIN,
            relief="flat",
            highlightthickness=1,
            highlightbackground=self.BORDER
        )
        self.pigment_text.grid(row=0, column=0, sticky="nsew")
        self.pigment_text.config(state="disabled")
        
        self.pigment_scroll = ttk.Scrollbar(self.text_scroll_frame, orient="vertical", command=self.pigment_text.yview)
        self.pigment_scroll.grid(row=0, column=1, sticky="ns")
        self.pigment_text.configure(yscrollcommand=self.pigment_scroll.set)
        
        # Panel lateral de selección humana de elementos
        self.discern_frame = ttk.LabelFrame(self.pigment_frame, text="Discernimiento Humano (Elementos)", padding=5)
        self.discern_frame.grid(row=0, column=1, sticky="nsew")
        self.discern_frame.rowconfigure(0, weight=1)
        self.discern_frame.columnconfigure(0, weight=1)
        
        self.discern_canvas = tk.Canvas(self.discern_frame, bg=self.BG_PANEL, bd=0, highlightthickness=0)
        self.discern_scroll = ttk.Scrollbar(self.discern_frame, orient="vertical", command=self.discern_canvas.yview)
        self.discern_scroll.pack(side="right", fill="y")
        self.discern_canvas.pack(side="left", fill="both", expand=True)
        self.discern_canvas.configure(yscrollcommand=self.discern_scroll.set)
        
        self.discern_inner = ttk.Frame(self.discern_canvas)
        self.discern_canvas_window = self.discern_canvas.create_window((0, 0), window=self.discern_inner, anchor="nw")
        
        def on_frame_configure(e):
            self.discern_canvas.configure(scrollregion=self.discern_canvas.bbox("all"))
        self.discern_inner.bind("<Configure>", on_frame_configure)
        
        # Elementos disponibles para discernir
        self.elementos_discern_vars = {}
        elementos_comunes = sorted(list(set(key.split('_')[0] for key in lectura_espectros.XRF_ELEMENTS.keys())))
        
        for idx, el in enumerate(elementos_comunes):
            var = tk.BooleanVar(value=False)
            self.elementos_discern_vars[el] = var
            cb = ttk.Checkbutton(
                self.discern_inner, 
                text=el, 
                variable=var, 
                command=self.on_discern_checkbox_changed
            )
            row = idx // 2
            col = idx % 2
            cb.grid(row=row, column=col, sticky="w", padx=5, pady=2)

        # Pestaña 3: Agrupamiento y Huella de Pigmentos (PCA)
        self.tab_pca = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_pca, text="Agrupamiento (PCA)")
        self.tab_pca.rowconfigure(2, weight=1) # El canvas ahora está en la fila 2 para dar espacio a la barra
        self.tab_pca.columnconfigure(0, weight=1)
        
        lbl_info_pca = ttk.Label(
            self.tab_pca,
            text="Agrupamiento Químico Masivo (PCA de todos los espectros en base a su perfil XRF):",
            font=("Helvetica", 10, "bold"),
            padding=10
        )
        lbl_info_pca.grid(row=0, column=0, sticky="w")

        # Barra de Herramientas de Navegación de Matplotlib para Zoom/Pan en el PCA
        self.toolbar_frame_pca = ttk.Frame(self.tab_pca)
        self.toolbar_frame_pca.grid(row=1, column=0, sticky="ew", padx=10)

        # Gráfica de PCA
        self.fig_pca = Figure(figsize=(7, 5), dpi=100, facecolor=self.BG_PANEL)
        self.ax_pca = self.fig_pca.add_subplot(111)
        self.ax_pca.set_facecolor(self.BG_CARD)
        self.ax_pca.set_title("PCA - Huella Química de los Espectros", color=self.FG_MAIN, weight='bold')
        self.ax_pca.set_xlabel("PC1", color=self.FG_MUTED)
        self.ax_pca.set_ylabel("PC2", color=self.FG_MUTED)
        self.ax_pca.tick_params(colors=self.FG_MUTED, which='both')
        for spine in self.ax_pca.spines.values():
            spine.set_color(self.BORDER)
        self.ax_pca.grid(True, linestyle="--", alpha=0.15, color=self.FG_MUTED)

        self.canvas_pca = FigureCanvasTkAgg(self.fig_pca, master=self.tab_pca)
        self.canvas_pca.get_tk_widget().grid(row=2, column=0, sticky="nsew", padx=10, pady=(0, 10))

        # Registrar la barra de herramientas
        self.toolbar_pca = NavigationToolbar2Tk(self.canvas_pca, self.toolbar_frame_pca)
        self.toolbar_pca.update()

        # Conectar eventos de interacción
        self.canvas_pca.mpl_connect("motion_notify_event", self.hover_pca)
        self.canvas_pca.mpl_connect("pick_event", self.on_pick_pca)

        # Controles de PCA y Exportación para PAST
        self.frame_controles_pca = ttk.LabelFrame(self.tab_pca, text=" Configuración de PCA y Exportación PAST ", padding=10)
        self.frame_controles_pca.grid(row=3, column=0, sticky="ew", padx=10, pady=(0, 10))
        
        lbl_roi = ttk.Label(self.frame_controles_pca, text="Rango Canales (ROI):")
        lbl_roi.grid(row=0, column=0, sticky="w", padx=5, pady=2)
        
        self.entry_canal_min = ttk.Entry(self.frame_controles_pca, width=6)
        self.entry_canal_min.insert(0, "100")
        self.entry_canal_min.grid(row=0, column=1, sticky="w", padx=2, pady=2)
        
        lbl_a = ttk.Label(self.frame_controles_pca, text="a")
        lbl_a.grid(row=0, column=2, padx=2)
        
        self.entry_canal_max = ttk.Entry(self.frame_controles_pca, width=6)
        self.entry_canal_max.insert(0, "1000")
        self.entry_canal_max.grid(row=0, column=3, sticky="w", padx=2, pady=2)
        
        lbl_metodo = ttk.Label(self.frame_controles_pca, text="Método:")
        lbl_metodo.grid(row=0, column=4, sticky="w", padx=(15, 5), pady=2)
        
        self.combo_metodo = ttk.Combobox(self.frame_controles_pca, values=["Covarianza (Recomendado)", "Correlación"], state="readonly", width=22)
        self.combo_metodo.current(0)
        self.combo_metodo.grid(row=0, column=5, sticky="w", padx=2, pady=2)
        
        self.var_alinear_signos = tk.BooleanVar(value=True)
        self.chk_alinear = ttk.Checkbutton(self.frame_controles_pca, text="Alinear Signos SVD", variable=self.var_alinear_signos)
        self.chk_alinear.grid(row=0, column=6, sticky="w", padx=(15, 5), pady=2)
        
        self.btn_recalcular_pca = ttk.Button(
            self.frame_controles_pca, 
            text="🔄 Recalcular PCA", 
            command=self.graficar_pca,
            style='Action.TButton'
        )
        self.btn_recalcular_pca.grid(row=1, column=0, columnspan=2, sticky="ew", padx=5, pady=5)
        
        self.btn_exportar_past_esp = ttk.Button(
            self.frame_controles_pca,
            text="📥 Exportar PAST (Espectros)",
            command=self.exportar_past_espectros,
            style='Action.TButton'
        )
        self.btn_exportar_past_esp.grid(row=1, column=4, columnspan=2, sticky="ew", padx=5, pady=5)
        
        self.btn_exportar_past_elem = ttk.Button(
            self.frame_controles_pca,
            text="📥 Exportar PAST (Elementos)",
            command=self.exportar_past_elementos,
            style='Action.TButton'
        )
        self.btn_exportar_past_elem.grid(row=1, column=6, sticky="ew", padx=5, pady=5)

        # Pestaña 4: Guía de Pigmentos (Teoría y Referencia)
        self.tab_guia = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_guia, text="Guía de Pigmentos")
        self.tab_guia.rowconfigure(0, weight=1)
        self.tab_guia.columnconfigure(0, weight=1)
        
        # Cuadro de texto scrollable para mostrar la guía de pigmentos y sugerencias
        self.guia_text = tk.Text(
            self.tab_guia, 
            font=("Consolas", 10) if os.name != 'nt' else ("Courier New", 10), 
            wrap="word", 
            bg=self.BG_CARD, 
            fg=self.FG_MAIN, 
            insertbackground=self.FG_MAIN,
            relief="flat",
            highlightthickness=1,
            highlightbackground=self.BORDER
        )
        self.guia_text.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        
        self.guia_scroll = ttk.Scrollbar(self.tab_guia, orient="vertical", command=self.guia_text.yview)
        self.guia_scroll.grid(row=0, column=1, sticky="ns", pady=10)
        self.guia_text.configure(yscrollcommand=self.guia_scroll.set)
        
        # Cargar contenido de la guía
        self.popular_guia_teorica()
        self.configurar_tooltips()

    def configurar_eventos(self):
        self.file_tree.bind("<<TreeviewSelect>>", self.on_tree_select)
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)
        self.file_tree.bind("<Button-2>", self.mostrar_menu_contextual)
        self.file_tree.bind("<Button-3>", self.mostrar_menu_contextual)

    def configurar_tooltips(self):
        Tooltip(self.drop_label, "Zona de arrastre: Suelta aquí archivos .pdz o proyectos .rtx/.rrtx para cargarlos.")
        Tooltip(self.btn_buscar, "Cargar archivos XRF .pdz o proyectos Artax (.rtx/.rrtx) de forma manual.")
        Tooltip(self.btn_buscar_carpeta, "Cargar una carpeta completa que contenga espectros XRF.")
        Tooltip(self.search_entry, "Buscar muestras por nombre o ID dentro del árbol de archivos.")
        Tooltip(self.btn_crear_carpeta, "Crear una nueva subcarpeta en la ubicación actual del árbol.")
        Tooltip(self.btn_mover_a, "Mover los archivos o carpetas seleccionadas a otra subcarpeta.")
        Tooltip(self.btn_theme, "Cambiar entre modo nocturno (tema oscuro) y modo de impresión (tema claro).")
        Tooltip(self.combo_rango, "Seleccionar rango de energía máximo (16, 30, 40 keV o ajuste manual).")
        Tooltip(self.entry_rango_manual, "Escribir valor de energía máximo personalizado en keV (presiona Enter).")
        Tooltip(self.chk_etiquetas_picos, "Mostrar u ocultar los nombres de los elementos identificados sobre los picos en la gráfica.")
        Tooltip(self.btn_procesar, "Alternar el cálculo automático del fondo continuo y la identificación de picos.")
        Tooltip(self.btn_excel, "Exportar los datos consolidados y resúmenes de todas las muestras a un archivo Excel.")
        Tooltip(self.btn_excel_graficas, "Exportar reporte detallado en Excel con hojas individuales y gráficas.")
        Tooltip(self.btn_guardar_rtx, "Guardar la sesión y avances del proyecto en un archivo XML .rtx compatible con Bruker ARTAX.")
        Tooltip(self.btn_artax_filtrado, "Exportar reporte con formato Bruker ARTAX (todos.xls) conteniendo todos los espectros individuales, pero limitado únicamente a las líneas de elementos marcadas en el panel de referencias.")
        Tooltip(self.btn_copiar_portapapeles, "Copiar la imagen de la gráfica actual al portapapeles para pegarla (Ctrl+V) sin guardar archivo.")
        Tooltip(self.btn_limpiar, "Eliminar todos los espectros y carpetas cargadas en la aplicación.")

    def actualizar_panel_referencia(self, elementos_detectados=None):
        """
        Reconstruye la lista de checkboxes de líneas de referencia.
        Si se le pasan elementos detectados, muestra solo los detectados y los asociados
        a patrimonio cultural / metales.
        Evita la recreación de widgets si la lista mostrada ya es idéntica (elimina el parpadeo).
        """
        elementos_todos = sorted(list(set(key.split('_')[0] for key in lectura_espectros.XRF_ELEMENTS.keys())))
        
        if elementos_detectados and len(elementos_detectados) > 0:
            detectados_set, asociados_set, relevantes_set = obtener_elementos_relevantes_patrimonio(elementos_detectados)
            elementos_a_mostrar = [el for el in elementos_todos if el in relevantes_set]
            header_text = f"Líneas de Referencia ({len(elementos_a_mostrar)} detectados/asociados)"
        else:
            detectados_set = set()
            asociados_set = set()
            elementos_a_mostrar = elementos_todos
            header_text = "Líneas de Referencia (Todos los elementos)"

        clave_actual = (tuple(elementos_a_mostrar), tuple(sorted(detectados_set)))
        if getattr(self, '_ref_panel_rendered_key', None) == clave_actual:
            return

        self._ref_panel_rendered_key = clave_actual
        self.ref_frame.config(text=header_text)

        # Limpiar widgets actuales en ref_frame
        for widget in self.ref_frame.winfo_children():
            widget.destroy()

        num_cols = 4
        for c in range(num_cols):
            self.ref_frame.columnconfigure(c, weight=1, uniform="ref_col")

        for i, el in enumerate(elementos_a_mostrar):
            if el not in self.elementos_var:
                self.elementos_var[el] = tk.BooleanVar(value=False)
            
            if el in detectados_set:
                txt = f"🟢 {el}"
            elif el in asociados_set:
                txt = f"🟡 {el}"
            else:
                txt = el

            cb = ttk.Checkbutton(self.ref_frame, text=txt, variable=self.elementos_var[el], command=self.on_element_toggle)
            row = i // num_cols
            col = i % num_cols
            cb.grid(row=row, column=col, sticky="ew", padx=3, pady=2)

    def obtener_rango_max(self):
        """
        Retorna el límite máximo del rango de energía en keV.
        Soporta opciones predefinidas ("16", "30", "40") o ajuste manual.
        """
        val = self.rango_max_var.get()
        if val == "Manual":
            try:
                v = float(self.rango_manual_var.get())
                return max(1.0, min(100.0, v))
            except (AttributeError, ValueError):
                return 16.0
        else:
            try:
                return float(val)
            except (AttributeError, ValueError):
                return 16.0

    def on_rango_changed(self, event=None):
        """
        Maneja el cambio en el selector de rango de energía (16, 30, 40 o Manual).
        """
        if self.rango_max_var.get() == "Manual":
            self.entry_rango_manual.grid(row=0, column=2, sticky="w", padx=(2, 5))
            self.entry_rango_manual.focus_set()
        else:
            self.entry_rango_manual.grid_remove()
        
        self.on_element_toggle()

    def on_element_toggle(self):
        # Redibujar la gráfica para incluir/quitar líneas de referencia
        self.on_tree_select(None)

    def on_tab_changed(self, event):
        # Si entramos a la pestaña de PCA, refrescar automáticamente
        selected_tab = self.notebook.index(self.notebook.select())
        if selected_tab == 2:
            self.graficar_pca()

    def toggle_procesamiento(self):
        self.mostrar_fondo_y_picos = not self.mostrar_fondo_y_picos
        if self.mostrar_fondo_y_picos:
            self.btn_procesar.config(text="✓ Análisis Científico (Activo)", style='Success.TButton')
            messagebox.showinfo("Análisis Activo", 
                                "Cálculo de fondo SNIP e identificación de picos activado.\n"
                                "Haz clic en cualquier espectro en el árbol. La pestaña 'Análisis Cuantitativo' se rellenará automáticamente.")
        else:
            self.btn_procesar.config(text="⚙️ Calcular Picos y Fondo", style='Action.TButton')
        
        self.on_tree_select(None)

    def limpiar_datos(self):
        self._ref_panel_rendered_key = None
        self.espectros_datos.clear()
        for item in self.file_tree.get_children():
            self.file_tree.delete(item)
            
        # Limpiar tablas y gráficas
        self.ax.clear()
        self.ax.set_facecolor(self.BG_CARD)
        self.ax.set_title("Espectro XRF", color=self.FG_MAIN, weight='bold')
        self.ax.set_xlabel("Energía (keV)", color=self.FG_MUTED)
        self.ax.set_ylabel("Cuentas", color=self.FG_MUTED)
        self.ax.tick_params(colors=self.FG_MUTED, which='both')
        for spine in self.ax.spines.values():
            spine.set_color(self.BORDER)
        self.ax.grid(True, linestyle="--", alpha=0.15, color=self.FG_MUTED)
        self.canvas.draw()
        
        self.ax_pca.clear()
        self.ax_pca.set_facecolor(self.BG_CARD)
        self.ax_pca.set_title("PCA - Huella Química de los Espectros", color=self.FG_MAIN, weight='bold')
        self.ax_pca.set_xlabel("PC1", color=self.FG_MUTED)
        self.ax_pca.set_ylabel("PC2", color=self.FG_MUTED)
        self.ax_pca.tick_params(colors=self.FG_MUTED, which='both')
        for spine in self.ax_pca.spines.values():
            spine.set_color(self.BORDER)
        self.ax_pca.grid(True, linestyle="--", alpha=0.15, color=self.FG_MUTED)
        self.canvas_pca.draw()
        
        for item in self.peaks_table.get_children():
            self.peaks_table.delete(item)
        
        if hasattr(self, 'pigment_text'):
            self.pigment_text.config(state="normal")
            self.pigment_text.delete("1.0", tk.END)
            self.pigment_text.config(state="disabled")
            
        self.drop_label.config(bg=self.BG_CARD, fg=self.FG_MUTED, text="Arrastra y suelta aquí tus archivos\n.pdz o tu proyecto .rtx de Artax")

    def agregar_al_arbol(self, categorias, nombre_mostrar, datos_espectro=None):
        parent = ""
        for cat in categorias:
            cat_id = cat if parent == "" else f"{parent}/{cat}"
            if not self.file_tree.exists(cat_id):
                cat_visible = cat.replace("Points ", "")
                self.file_tree.insert(parent, "end", cat_id, text=f"📁 {cat_visible}", open=True)
            parent = cat_id
            
        nombre_limpio = nombre_mostrar.split('@')[0] if '@' in nombre_mostrar else nombre_mostrar
        esp_id = f"{parent}/{nombre_mostrar}" if parent else nombre_mostrar
        
        if not self.file_tree.exists(esp_id):
            self.file_tree.insert(parent, "end", esp_id, text=f"📊 {nombre_limpio}")
            if datos_espectro:
                # Almacenar datos y también mapear su grupo padre para PCA
                datos_espectro['grupo_padre'] = categorias[-1] if categorias else 'Sin clasificar'
                self.espectros_datos[esp_id] = datos_espectro

    def procesar_drop(self, event):
        rutas_limpias = self.root.tk.splitlist(event.data)
        self.cargar_archivos(rutas_limpias)

    def seleccionar_archivos(self):
        rutas = filedialog.askopenfilenames(
            title="Seleccionar archivos XRF o Proyecto Artax",
            filetypes=[
                ("Archivos XRF", ("*.rtx", "*.rrtx", "*.pdz")),
                ("Proyecto Artax", ("*.rtx", "*.rrtx")),
                ("Espectros Bruker", "*.pdz"),
                ("Todos los archivos", "*")
            ]
        )
        if rutas:
            self.cargar_archivos(list(rutas))

    def seleccionar_carpeta(self):
        carpeta = filedialog.askdirectory(
            title="Seleccionar Carpeta con Espectros XRF"
        )
        if carpeta:
            self.cargar_archivos([carpeta])

    def cargar_archivos(self, rutas_limpias):
        if hasattr(self, 'search_var'):
            self.search_var.set("")
        # Expandir directorios recursivamente
        archivos_finales = []
        origen_raiz = {}
        
        for r in rutas_limpias:
            if os.path.isdir(r):
                for root_dir, dirs, files in os.walk(r):
                    for f in files:
                        full_p = os.path.join(root_dir, f)
                        archivos_finales.append(full_p)
                        origen_raiz[full_p] = r
            else:
                archivos_finales.append(r)
                origen_raiz[r] = os.path.dirname(r)
                
        rtx_files = [r for r in archivos_finales if r.lower().endswith('.rtx') or r.lower().endswith('.rrtx')]
        pdz_files = [r for r in archivos_finales if r.lower().endswith('.pdz')]
        
        total_nuevos = 0
        
        # 1. Cargar Proyecto RTX
        for rtx_path in rtx_files:
            dir_rtx = os.path.dirname(rtx_path)
            # Buscar calibración en la carpeta del proyecto y sus padres (hasta 3 niveles)
            calib = lectura_espectros.obtener_calibracion(dir_rtx)
            if not calib:
                curr_c = dir_rtx
                for _ in range(3):
                    parent_c = os.path.dirname(curr_c)
                    if not parent_c or parent_c == curr_c:
                        break
                    calib = lectura_espectros.obtener_calibracion(parent_c)
                    if calib:
                        break
                    curr_c = parent_c
            if calib:
                print(f"[Calibración] Encontrada para proyecto en {dir_rtx}: pendiente={calib[0]:.6f}, intercepto={calib[1]:.6f}")
            try:
                # Pre-escanear y cachear todos los archivos .pdz en el directorio del RTX
                # y en sus directorios superiores (hasta 3 niveles arriba), omitiendo venv/git.
                pdz_cache = {}
                directorios_busqueda = [dir_rtx]
                curr_dir = dir_rtx
                for _ in range(3):
                    parent_dir = os.path.dirname(curr_dir)
                    if not parent_dir or parent_dir == curr_dir:
                        break
                    directorios_busqueda.append(parent_dir)
                    curr_dir = parent_dir
                
                # Escanear de arriba hacia abajo (niveles más genéricos primero, para que los específicos reescriban)
                for d in reversed(directorios_busqueda):
                    if not os.path.isdir(d):
                        continue
                    for root_d, dirs_d, files_d in os.walk(d):
                        # Podar directorios no deseados in-place para que os.walk no entre en ellos
                        dirs_d[:] = [dirname for dirname in dirs_d if dirname not in ('venv', '.git', '__pycache__', 'node_modules')]
                        for f in files_d:
                            if f.lower().endswith('.pdz'):
                                pdz_cache[f.lower()] = os.path.join(root_d, f)

                espectros = lectura_espectros.parsear_rtx(rtx_path)
                for esp in espectros:
                    pdz_name = esp['archivo_pdz']
                    pdz_full_path = os.path.join(dir_rtx, pdz_name)
                    
                    if not os.path.exists(pdz_full_path):
                        encontrado_ruta = None
                        if pdz_name.lower() in pdz_cache:
                            encontrado_ruta = pdz_cache[pdz_name.lower()]
                        else:
                            # Búsqueda parcial de respaldo si no hay coincidencia exacta
                            for f_lower, full_path in pdz_cache.items():
                                if f_lower in pdz_name.lower() or pdz_name.lower() in f_lower:
                                    encontrado_ruta = full_path
                                    break
                        if encontrado_ruta:
                            pdz_full_path = encontrado_ruta
                            
                    datos_esp = None
                    if os.path.exists(pdz_full_path):
                        try:
                            datos_esp = lectura_espectros.leer_pdz(pdz_full_path, calibracion=calib)
                            total_nuevos += 1
                        except Exception as e:
                            print(f"Error cargando pdz: {pdz_name}: {e}")
                            
                    if datos_esp is None and 'xml_data' in esp:
                        try:
                            xml_meta = esp['xml_data']
                            counts = xml_meta['counts']
                            if counts:
                                if calib:
                                    slope, intercept = calib
                                else:
                                    slope, intercept = xml_meta['xml_calib']
                                    
                                energias = [i * slope + intercept for i in range(len(counts))]
                                df = pd.DataFrame({
                                    'Energia_keV': energias,
                                    'Cuentas': counts
                                })
                                datos_esp = {
                                    'metadata': {
                                        'num_channels': xml_meta['num_channels'],
                                        'ev_per_channel': xml_meta['ev_per_channel'],
                                        'live_time': xml_meta['live_time'],
                                        'xray_voltage_kv': xml_meta['xray_voltage_kv'],
                                        'xray_filament_current': xml_meta['xray_filament_current'],
                                        'version': None,
                                        'file_type': None,
                                        'nombre_archivo': pdz_name
                                    },
                                    'datos': df
                                }
                                total_nuevos += 1
                        except Exception as ex_fallback:
                            print(f"Error reconstruyendo espectro desde RTX: {pdz_name}: {ex_fallback}")
                    
                    self.agregar_al_arbol(
                        categorias=esp['categorias'],
                        nombre_mostrar=esp['nombre'],
                        datos_espectro=datos_esp
                    )
                self.drop_label.config(bg=self.SUCCESS, fg=self.FG_MAIN, text=f"¡Proyecto RTX cargado ({total_nuevos} espectros vinculados)!")
            except Exception as e:
                messagebox.showerror("Error", f"Error al parsear el archivo RTX:\n{e}")

        # 2. Cargar archivos .pdz sueltos
        pdz_sueltos = 0
        for pdz_path in pdz_files:
            nombre_pdz = os.path.basename(pdz_path)
            ya_existe = any(d.get('metadata', {}).get('nombre_archivo') == nombre_pdz for d in self.espectros_datos.values())
            
            if not ya_existe:
                try:
                    dir_pdz = os.path.dirname(pdz_path)
                    calib = lectura_espectros.obtener_calibracion(dir_pdz)
                    datos_esp = lectura_espectros.leer_pdz(pdz_path, calibracion=calib)
                    
                    # Determinar estructura de categorías basándose en la ruta relativa
                    raiz = origen_raiz.get(pdz_path, dir_pdz)
                    rel_dir = os.path.relpath(dir_pdz, raiz)
                    
                    if rel_dir == "." or rel_dir == "":
                        categorias = [os.path.basename(raiz)] if os.path.isdir(raiz) else ["Espectros sin clasificar"]
                    else:
                        categorias = [os.path.basename(raiz)] + rel_dir.split(os.sep)
                        
                    self.agregar_al_arbol(
                        categorias=categorias,
                        nombre_mostrar=nombre_pdz,
                        datos_espectro=datos_esp
                    )
                    pdz_sueltos += 1
                except Exception as e:
                    print(f"Error cargando pdz suelto: {nombre_pdz}: {e}")
                    
        if pdz_sueltos > 0:
            self.drop_label.config(bg=self.SUCCESS, fg=self.FG_MAIN, text=f"¡Se agregaron {pdz_sueltos} espectros sueltos!")
        self.actualizar_espectros_suma()

    def crear_nueva_carpeta(self):
        from tkinter import simpledialog
        nombre = simpledialog.askstring("Nueva Carpeta", "Ingresa el nombre de la subcarpeta:", parent=self.root)
        if not nombre:
            return
            
        # Determinar el padre. Si hay algo seleccionado en el árbol, podemos crearlo como subcarpeta de eso (si es una carpeta),
        # o en la raíz si no hay selección o es un archivo.
        parent = ""
        selected = self.file_tree.selection()
        if selected:
            sel_id = selected[0]
            if sel_id not in self.espectros_datos:
                parent = sel_id
                
        # Generar ID único para la carpeta
        cat_id = nombre if parent == "" else f"{parent}/{nombre}"
        
        # Validar si ya existe
        if self.file_tree.exists(cat_id):
            messagebox.showwarning("Advertencia", f"La carpeta '{nombre}' ya existe en esta ubicación.")
            return
            
        self.file_tree.insert(parent, "end", cat_id, text=f"📁 {nombre}", open=True)

    def obtener_carpetas_arbol(self):
        carpetas = ["(Raíz)"]
        
        def recorrer(item):
            for child in self.file_tree.get_children(item):
                if child not in self.espectros_datos:
                    carpetas.append(child)
                    recorrer(child)
                    
        recorrer("")
        return carpetas

    def actualizar_grupo_padre_recursivo(self, item_id):
        grupo_nombre = self.file_tree.item(item_id, "text").replace("📁 ", "")
        
        def recorrer(nodo, grupo):
            if nodo in self.espectros_datos:
                self.espectros_datos[nodo]['grupo_padre'] = grupo
            else:
                nombre_carpeta = self.file_tree.item(nodo, "text").replace("📁 ", "")
                for child in self.file_tree.get_children(nodo):
                    recorrer(child, nombre_carpeta)
                    
        recorrer(item_id, grupo_nombre)

    def mover_espectro_a(self):
        if hasattr(self, 'search_var'):
            self.search_var.set("")
        selected = self.file_tree.selection()
        if not selected:
            messagebox.showwarning("Advertencia", "Por favor, selecciona los archivos o carpetas que deseas mover.")
            return
            
        # Obtener todas las carpetas destino
        carpetas = self.obtener_carpetas_arbol()
        
        # Crear ventana emergente
        pop = tk.Toplevel(self.root)
        pop.title("Mover Elementos")
        pop.geometry("400x350")
        pop.transient(self.root)
        pop.grab_set()
        pop.configure(bg=self.BG_MAIN)
        
        # Centrar ventana emergente
        pop.update_idletasks()
        width = pop.winfo_width()
        height = pop.winfo_height()
        x = (pop.winfo_screenwidth() // 2) - (width // 2)
        y = (pop.winfo_screenheight() // 2) - (height // 2)
        pop.geometry(f"+{x}+{y}")
        
        ttk.Label(pop, text=f"Mover {len(selected)} elemento(s) a:", font=("Helvetica", 10, "bold")).pack(pady=10)
        
        # Listbox para las carpetas
        frame_list = ttk.Frame(pop)
        frame_list.pack(fill="both", expand=True, padx=15, pady=5)
        
        scroll = ttk.Scrollbar(frame_list, orient="vertical")
        scroll.pack(side="right", fill="y")
        
        listbox = tk.Listbox(
            frame_list, 
            yscrollcommand=scroll.set, 
            bg=self.BG_CARD, 
            fg=self.FG_MAIN, 
            selectbackground=self.ACCENT, 
            selectforeground=self.FG_MAIN,
            highlightcolor=self.BORDER,
            highlightbackground=self.BORDER,
            bd=0,
            font=("Helvetica", 10)
        )
        listbox.pack(fill="both", expand=True)
        scroll.config(command=listbox.yview)
        
        # Llenar listbox
        for c in carpetas:
            if c == "(Raíz)":
                mostrar = "📁 [Raíz]"
            else:
                mostrar = "📁 " + c
            listbox.insert("end", mostrar)
            
        def confirmar():
            sel_idx = listbox.curselection()
            if not sel_idx:
                messagebox.showwarning("Advertencia", "Selecciona una carpeta destino.", parent=pop)
                return
                
            dest_cat = carpetas[sel_idx[0]]
            dest_id = "" if dest_cat == "(Raíz)" else dest_cat
            
            # Validar que no se mueva una carpeta a sí misma o a una de sus subcarpetas
            for item_a_mover in selected:
                temp = dest_id
                while temp:
                    if temp == item_a_mover:
                        messagebox.showerror("Error", f"No puedes mover la carpeta '{self.file_tree.item(item_a_mover, 'text')}' dentro de sí misma o de sus subcarpetas.", parent=pop)
                        return
                    parts = temp.split('/')
                    if len(parts) > 1:
                        temp = "/".join(parts[:-1])
                    else:
                        temp = ""
            
            try:
                for item_a_mover in selected:
                    # Mover en el Treeview
                    self.file_tree.move(item_a_mover, dest_id, "end")
                    
                    # Actualizar el grupo padre en self.espectros_datos para las muestras movidas
                    es_muestra = item_a_mover in self.espectros_datos
                    if es_muestra:
                        grupo_nombre = self.file_tree.item(dest_id, "text").replace("📁 ", "") if dest_id else 'Sin clasificar'
                        self.espectros_datos[item_a_mover]['grupo_padre'] = grupo_nombre
                    else:
                        self.actualizar_grupo_padre_recursivo(item_a_mover)
                
                # Refrescar sumas, selección y gráficas
                self.actualizar_espectros_suma()
                self.on_tree_select(None)
                pop.destroy()
                messagebox.showinfo("Éxito", f"Se movieron {len(selected)} elemento(s) correctamente.")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudieron mover los elementos:\n{e}", parent=pop)
                
        # Botón de confirmación y cancelación
        btn_frame = ttk.Frame(pop, padding=10)
        btn_frame.pack(fill="x", side="bottom")
        
        btn_cancel = ttk.Button(btn_frame, text="Cancelar", command=pop.destroy, style='Danger.TButton')
        btn_cancel.pack(side="left", padx=5, expand=True, fill="x")
        
        btn_ok = ttk.Button(btn_frame, text="Mover", command=confirmar, style='Success.TButton')
        btn_ok.pack(side="right", padx=5, expand=True, fill="x")

    def mostrar_menu_contextual(self, event):
        item = self.file_tree.identify_row(event.y)
        if item:
            current_selection = self.file_tree.selection()
            if item not in current_selection:
                self.file_tree.selection_set(item)
            self.menu_contextual.post(event.x_root, event.y_root)

    def eliminar_elemento_arbol(self):
        if hasattr(self, 'search_var'):
            self.search_var.set("")
        selected = self.file_tree.selection()
        if not selected:
            return
            
        confirmar = messagebox.askyesno(
            "Confirmar Eliminación", 
            f"¿Estás seguro de que deseas eliminar los {len(selected)} elemento(s) seleccionado(s)?\n(Esto no eliminará los archivos físicos de tu disco)",
            parent=self.root
        )
        if not confirmar:
            return
            
        def eliminar_datos_recursivo(nodo):
            if nodo in self.espectros_datos:
                del self.espectros_datos[nodo]
            else:
                for child in self.file_tree.get_children(nodo):
                    eliminar_datos_recursivo(child)
                    
        for item_id in selected:
            eliminar_datos_recursivo(item_id)
            self.file_tree.delete(item_id)
            
        self.actualizar_espectros_suma()
        self.on_tree_select(None)

    def on_discern_checkbox_changed(self):
        selected = self.file_tree.selection()
        if selected and selected[0] in self.espectros_datos:
            item_id = selected[0]
            # Actualizar en memoria
            self.espectros_datos[item_id]['elementos_discernidos'] = [
                el for el, var in self.elementos_discern_vars.items() if var.get()
            ]
        # Re-evaluar e imprimir en self.pigment_text
        self.actualizar_sugerencias_discernidas()

    def actualizar_sugerencias_discernidas(self):
        elementos_elegidos = []
        for el, var in self.elementos_discern_vars.items():
            if var.get():
                elementos_elegidos.append(el)
                
        # Generar sugerencias basadas en el discernimiento del usuario
        sugerencias = lectura_espectros.sugerir_pigmentos(elementos_elegidos)
        
        self.pigment_text.config(state="normal")
        self.pigment_text.delete("1.0", tk.END)
        
        self.pigment_text.insert(tk.END, "Interpretación de Pigmentos por Discernimiento Humano:\n")
        self.pigment_text.insert(tk.END, f"Elementos seleccionados para análisis: {', '.join(elementos_elegidos) if elementos_elegidos else 'Ninguno'}\n\n")
        
        if sugerencias:
            for sug in sugerencias:
                self.pigment_text.insert(tk.END, f"• {sug['pigmento']} ({sug['color']}): {sug['justificacion']}\n")
        else:
            self.pigment_text.insert(tk.END, "No se encontraron correlaciones de pigmentos para los elementos seleccionados.\n")
            
        self.pigment_text.config(state="disabled")

    def on_search_change(self):
        query = self.search_var.get().lower().strip()
        
        if not hasattr(self, '_original_tree_structure'):
            self._original_tree_structure = {}
            
        if not query:
            self.restaurar_estructura_original()
            return
            
        if not self._original_tree_structure:
            self.guardar_estructura_actual()
            
        coinciden = set()
        
        # Encontrar qué muestras coinciden
        for item in self.espectros_datos:
            text = self.file_tree.item(item, "text").lower()
            if query in text:
                coinciden.add(item)
                # Asegurar de mostrar todos los ancestros
                temp = item
                while True:
                    parent = self._original_tree_structure.get(temp, {}).get('parent', '')
                    if not parent:
                        break
                    coinciden.add(parent)
                    temp = parent
                    
        # Encontrar qué carpetas coinciden por su propio nombre
        all_items = list(self._original_tree_structure.keys())
        for item in all_items:
            if item not in self.espectros_datos:
                text = self.file_tree.item(item, "text").lower()
                if query in text:
                    coinciden.add(item)
                    # Mostrar recursivamente todos los hijos de la carpeta
                    def agregar_hijos(nodo):
                        for child, info in self._original_tree_structure.items():
                            if info['parent'] == nodo:
                                coinciden.add(child)
                                agregar_hijos(child)
                    agregar_hijos(item)
                    # Y mostrar ancestros
                    temp = item
                    while True:
                        parent = self._original_tree_structure.get(temp, {}).get('parent', '')
                        if not parent:
                            break
                        coinciden.add(parent)
                        temp = parent
                        
        # Ocultar o mostrar de arriba a abajo
        for item in all_items:
            if item in coinciden:
                parent = self._original_tree_structure[item]['parent']
                index = self._original_tree_structure[item]['index']
                try:
                    self.file_tree.move(item, parent, index)
                    if item not in self.espectros_datos:
                        self.file_tree.item(item, open=True)
                except Exception:
                    pass
            else:
                try:
                    self.file_tree.detach(item)
                except Exception:
                    pass

    def guardar_estructura_actual(self):
        self._original_tree_structure = {}
        def recorrer(item):
            for idx, child in enumerate(self.file_tree.get_children(item)):
                self._original_tree_structure[child] = {
                    'parent': item,
                    'index': idx
                }
                recorrer(child)
        recorrer("")

    def restaurar_estructura_original(self):
        if not hasattr(self, '_original_tree_structure') or not self._original_tree_structure:
            return
        sorted_items = sorted(self._original_tree_structure.items(), key=lambda x: len(x[0].split('/')))
        for item, info in sorted_items:
            try:
                self.file_tree.move(item, info['parent'], info['index'])
            except Exception:
                pass
        self._original_tree_structure = {}

    def actualizar_espectros_suma(self):
        # 1. Identificar y eliminar nodos SUMA_AUTO existentes para evitar duplicación o sumas obsoletas
        nodos_a_eliminar = []
        def buscar_sumas(item):
            for child in self.file_tree.get_children(item):
                if child.endswith("/SUMA_AUTO"):
                    nodos_a_eliminar.append(child)
                else:
                    buscar_sumas(child)
        buscar_sumas("")
        
        for nodo in nodos_a_eliminar:
            if self.file_tree.exists(nodo):
                self.file_tree.delete(nodo)
            if nodo in self.espectros_datos:
                del self.espectros_datos[nodo]
                
        # 2. Encontrar todas las carpetas en el árbol
        carpetas = []
        def buscar_carpetas(item):
            for child in self.file_tree.get_children(item):
                if child not in self.espectros_datos:
                    carpetas.append(child)
                    buscar_carpetas(child)
        buscar_carpetas("")
        
        # 3. Para cada carpeta, calcular la suma de todos los espectros que contiene
        for folder_id in carpetas:
            espectros_reales = []
            todos = self.obtener_espectros_recursivos(folder_id)
            for name, esp in todos:
                if not esp.get('metadata', {}).get('es_suma', False):
                    espectros_reales.append(esp)
                    
            if len(espectros_reales) >= 2:
                # Sumar counts canal por canal
                base_df = espectros_reales[0]['datos']
                sum_counts = np.zeros(len(base_df))
                for esp in espectros_reales:
                    counts = esp['datos']['Cuentas'].values
                    if len(counts) == len(sum_counts):
                        sum_counts += counts
                    else:
                        min_len = min(len(counts), len(sum_counts))
                        sum_counts[:min_len] += counts[:min_len]
                        
                df_sum = base_df.copy()
                df_sum['Cuentas'] = sum_counts
                
                base_meta = espectros_reales[0]['metadata']
                total_live_time = sum(esp['metadata'].get('live_time', 0.0) for esp in espectros_reales)
                folder_name = os.path.basename(folder_id)
                
                metadata_sum = {
                    'nombre_archivo': f"SUMA_{folder_name}.pdz",
                    'live_time': total_live_time,
                    'xray_voltage_kv': base_meta.get('xray_voltage_kv', 0.0),
                    'xray_filament_current': base_meta.get('xray_filament_current', 0.0),
                    'es_suma': True
                }
                
                esp_id = f"{folder_id}/SUMA_AUTO"
                self.file_tree.insert(folder_id, "end", esp_id, text=f"📊 SUMA - {folder_name}")
                self.espectros_datos[esp_id] = {
                    'datos': df_sum,
                    'metadata': metadata_sum,
                    'grupo_padre': folder_name
                }

    def obtener_espectros_recursivos(self, item_id):
        espectros = []
        if item_id in self.espectros_datos:
            espectros.append((self.file_tree.item(item_id, "text"), self.espectros_datos[item_id]))
        else:
            children = self.file_tree.get_children(item_id)
            for child in children:
                espectros.extend(self.obtener_espectros_recursivos(child))
        return espectros

    def on_tree_select(self, event):
        selected_items = self.file_tree.selection()
        if not selected_items:
            return
            
        item_id = selected_items[0]
        espectros_a_graficar = self.obtener_espectros_recursivos(item_id)
        
        self.ax.clear()
        
        # Limpiar tabla cuantitativa y texto de pigmentos por defecto
        for item in self.peaks_table.get_children():
            self.peaks_table.delete(item)
            
        if hasattr(self, 'pigment_text'):
            self.pigment_text.config(state="normal")
            self.pigment_text.delete("1.0", tk.END)
            self.pigment_text.config(state="disabled")
            
        if not espectros_a_graficar:
            self.ax.set_title("Espectro XRF")
            self.ax.set_xlabel("Energía (keV)")
            self.ax.set_ylabel("Cuentas")
            self.ax.grid(True, linestyle="--", alpha=0.5)
            self.canvas.draw()
            return
            
        # Caso A: Gráfica individual detallada con Análisis Científico
        if len(espectros_a_graficar) == 1 and self.mostrar_fondo_y_picos:
            nombre, esp = espectros_a_graficar[0]
            df = esp['datos']
            meta = esp['metadata']
            
            counts = df['Cuentas'].values
            energias = df['Energia_keV'].values
            
            # Calcular fondo SNIP y buscar picos
            fondo = lectura_espectros.calcular_fondo_snip(counts)
            picos = lectura_espectros.buscar_picos(energias, counts, fondo)
            
            # Graficar bruto y fondo
            self.ax.plot(energias, counts, color='#1f77b4', label='Espectro Bruto', linewidth=1.2)
            self.ax.plot(energias, fondo, color='#ff7f0e', linestyle='--', label='Fondo (SNIP)', linewidth=1.2)
            
            # Dibujar y etiquetar picos
            for p in picos:
                self.ax.plot(p['energia_kev'], p['cuentas_brutas'], 'ro', markersize=4)
                if p['elemento'] and self.mostrar_etiquetas_picos_var.get():
                    self.ax.text(
                        p['energia_kev'], 
                        p['cuentas_brutas'] + (max(counts) * 0.02), 
                        p['elemento'], 
                        fontsize=8, 
                        rotation=90, 
                        verticalalignment='bottom', 
                        horizontalalignment='center',
                        color='#d62728',
                        weight='bold'
                    )
                    
            # Rellenar la tabla cuantitativa y recopilar elementos detectados
            elementos_detectados = []
            for p in picos:
                elem_display = p['elemento'] if p['elemento'] else 'Desconocido'
                if p['elemento']:
                    elementos_detectados.append((p['elemento'], p['area_relativa']))
                self.peaks_table.insert('', 'end', values=(
                    elem_display,
                    f"{p['energia_kev']:.3f}",
                    int(p['cuentas_brutas']),
                    f"{p['cuentas_netas']:.1f}",
                    f"{p['area_neta']:.1f}",
                    f"{p['area_relativa']:.2f}%"
                ))
                
            # Cargar elementos discernidos de esta muestra
            discernidos = esp.get('elementos_discernidos')
            if discernidos is None:
                # Si no está inicializado, extraer del análisis automático de picos
                elementos_lista = []
                for p in picos:
                    if p['elemento']:
                        el = p['elemento']
                        sym = el.split('(')[1].split(')')[0].strip() if '(' in el else str(el).strip()
                        elementos_lista.append(sym)
                discernidos = list(set(elementos_lista))
                esp['elementos_discernidos'] = discernidos
                
            # Actualizar el checklist de discernimiento humano
            for el, var in self.elementos_discern_vars.items():
                var.set(el in discernidos)
                
            # Mostrar las sugerencias basadas en el discernimiento actual
            self.actualizar_sugerencias_discernidas()
            
            # Actualizar el panel de líneas de referencia con sólo los elementos detectados y sus asociados
            self.actualizar_panel_referencia(discernidos)

            self.ax.set_title(f"Espectro: {meta['nombre_archivo']} (V: {meta['xray_voltage_kv']}kV, C: {meta['xray_filament_current']}uA)")
            self.ax.legend(fontsize=9, loc='upper right')
            
        # Caso B: Gráfica comparativa o individual sin análisis de fondo
        else:
            self.actualizar_panel_referencia(None)
            if hasattr(self, 'elementos_discern_vars'):
                for var in self.elementos_discern_vars.values():
                    var.set(False)
            max_legend = 12
            for nombre, esp in espectros_a_graficar:
                df = esp['datos']
                label = nombre.replace("📊 ", "")
                if len(espectros_a_graficar) > max_legend:
                    self.ax.plot(df['Energia_keV'], df['Cuentas'], linewidth=0.8, alpha=0.5)
                else:
                    self.ax.plot(df['Energia_keV'], df['Cuentas'], linewidth=1.0, label=label, alpha=0.8)
            
            nombre_nodo = self.file_tree.item(item_id, "text").replace("📁 ", "").replace("📊 ", "")
            if len(espectros_a_graficar) == 1:
                meta = espectros_a_graficar[0][1]['metadata']
                self.ax.set_title(f"Espectro: {meta['nombre_archivo']} (T: {meta['live_time']:.1f}s)")
            else:
                self.ax.set_title(f"Comparación: {nombre_nodo} ({len(espectros_a_graficar)} espectros)")
                if len(espectros_a_graficar) <= max_legend:
                    self.ax.legend(fontsize=8, loc='upper right')

        # Obtener rango máximo de energía seleccionado (16, 30, 40 keV o manual)
        rango_max = self.obtener_rango_max()

        # DIBUJAR LÍNEAS DE REFERENCIA DE ELEMENTOS MARCADOS
        for el, var in self.elementos_var.items():
            if var.get():
                for key, val in lectura_espectros.XRF_ELEMENTS.items():
                    if key.startswith(f"{el}_") or key == el:
                        energy, line, label = val
                        # Solo dibujar si está en el rango visible
                        if energy <= rango_max:
                            self.ax.axvline(x=energy, color='gray', linestyle=':', linewidth=1.0)
                            self.ax.text(
                                energy, 
                                self.ax.get_ylim()[1] * 0.95, 
                                f"{el} {line}", 
                                color=self.FG_MUTED, 
                                fontsize=7, 
                                rotation=90, 
                                verticalalignment='top', 
                                horizontalalignment='right'
                            )
                    
        self.ax.set_xlabel("Energía (keV)", color=self.FG_MUTED)
        self.ax.set_ylabel("Cuentas", color=self.FG_MUTED)
        self.ax.grid(True, linestyle="--", alpha=0.15, color=self.FG_MUTED)
        self.ax.set_xlim(0, rango_max)
        
        # Apply dark theme styling to all plot elements
        self.ax.set_facecolor(self.BG_CARD)
        self.ax.tick_params(colors=self.FG_MUTED, which='both')
        self.ax.xaxis.label.set_color(self.FG_MUTED)
        self.ax.yaxis.label.set_color(self.FG_MUTED)
        self.ax.title.set_color(self.FG_MAIN)
        for spine in self.ax.spines.values():
            spine.set_color(self.BORDER)
            
        legend = self.ax.legend(fontsize=8, loc='upper right')
        if legend:
            legend.get_frame().set_facecolor(self.BG_PANEL)
            legend.get_frame().set_edgecolor(self.BORDER)
            for text in legend.get_texts():
                text.set_color(self.FG_MAIN)
                
        self.canvas.draw()

    def graficar_pca(self):
        # Filtrar espectros que realmente tengan datos cargados en memoria (excluyendo espectros suma)
        espectros_validos = {k: v for k, v in self.espectros_datos.items() if v.get('datos') is not None and not v.get('metadata', {}).get('es_suma', False)}
        
        if len(espectros_validos) < 3:
            self.ax_pca.clear()
            self.ax_pca.text(0.5, 0.5, "Carga un proyecto .rtx con archivos .pdz\n(mínimo 3 espectros) para ver el PCA.",
                             horizontalalignment='center', verticalalignment='center', transform=self.ax_pca.transAxes,
                             fontsize=11, style='italic', color=self.FG_MUTED)
            self.ax_pca.set_facecolor(self.BG_CARD)
            self.ax_pca.tick_params(colors=self.FG_MUTED, which='both')
            for spine in self.ax_pca.spines.values():
                spine.set_color(self.BORDER)
            self.canvas_pca.draw()
            return
            
        # NUEVOS PARÁMETROS LEÍDOS DE LA INTERFAZ
        try:
            c_min = int(self.entry_canal_min.get())
        except Exception:
            c_min = 100
            self.entry_canal_min.delete(0, tk.END)
            self.entry_canal_min.insert(0, "100")
            
        try:
            c_max = int(self.entry_canal_max.get())
        except Exception:
            c_max = 1000
            self.entry_canal_max.delete(0, tk.END)
            self.entry_canal_max.insert(0, "1000")
            
        metodo_combo = self.combo_metodo.get()
        metodo = "correlacion" if "Correlación" in metodo_combo else "covarianza"
        alinear = self.var_alinear_signos.get()
            
        res = lectura_espectros.calcular_pca_espectros(
            espectros_validos, 
            canal_min=c_min, 
            canal_max=c_max, 
            metodo=metodo, 
            alinear_signos=alinear
        )
        if not res:
            return
            
        self.ax_pca.clear()
        
        resultados = res['resultados']
        var_pc1 = res['var_pc1']
        var_pc2 = res['var_pc2']
        
        # Agrupar por categoría
        categories_dict = {}
        for r in resultados:
            cat = r['categoria']
            if cat not in categories_dict:
                categories_dict[cat] = {'x': [], 'y': [], 'names': [], 'item_ids': []}
            categories_dict[cat]['x'].append(r['pc1'])
            categories_dict[cat]['y'].append(r['pc2'])
            categories_dict[cat]['names'].append(r['nombre'])
            categories_dict[cat]['item_ids'].append(r['item_id'])
            
        # Tabla de colores por defecto si no se detecta color en el nombre
        import matplotlib.colors as mcolors
        colores_defecto = list(mcolors.TABLEAU_COLORS.values())
        
        # Mapear colores de los puntos según el color del pigmento descrito en la categoría
        def obtener_color_por_nombre(nombre_cat):
            cat_lower = nombre_cat.lower()
            if "rojo" in cat_lower and "negro" in cat_lower:
                return "#800f2f" # Rojo oscuro / mezcla
            elif "rojo" in cat_lower or "red" in cat_lower:
                return "#e63946" # Rojo premium
            elif "negro" in cat_lower or "black" in cat_lower:
                return "#343a40" # Negro / carbón (visible tanto en tema claro como oscuro)
            elif "ocre" in cat_lower or "ochre" in cat_lower:
                return "#b5842c" # Ocre premium
            elif "rosa" in cat_lower or "rose" in cat_lower or "pink" in cat_lower:
                return "#ff85a1" # Rosa premium
            elif "naranja" in cat_lower or "orange" in cat_lower:
                return "#ff6b35" # Naranja coral
            elif "azulverde" in cat_lower or "azul_verde" in cat_lower or "azul verde" in cat_lower:
                return "#00b4d8" # Azul-Verde / Turquesa
            elif "azul" in cat_lower or "blue" in cat_lower:
                return "#1f77b4" # Azul
            elif "verde" in cat_lower or "green" in cat_lower:
                return "#52b788" # Verde premium
            elif "amarillo" in cat_lower or "yellow" in cat_lower:
                return "#ffd166" # Amarillo cálido
            elif "blanco" in cat_lower or "white" in cat_lower:
                return "#f8f9fa" # Blanco
            elif "cafe" in cat_lower or "café" in cat_lower or "marron" in cat_lower or "marrón" in cat_lower or "brown" in cat_lower:
                return "#6f4e37" # Marrón
            elif "gris" in cat_lower or "gray" in cat_lower or "grey" in cat_lower:
                return "#6c757d" # Gris
            return None

        # Determinar el color del borde de los puntos según el tema activo para contraste
        borde_color = self.FG_MAIN if hasattr(self, 'FG_MAIN') else 'black'
        
        self.sc_plots = []
        for idx, (cat, pts) in enumerate(categories_dict.items()):
            color = obtener_color_por_nombre(cat)
            if not color:
                color = colores_defecto[idx % len(colores_defecto)]
                
            sc = self.ax_pca.scatter(pts['x'], pts['y'], label=cat, color=color, s=60, alpha=0.8, edgecolors=borde_color, picker=5)
            self.sc_plots.append((sc, pts))
            
        # Crear la anotación oculta para el hover (tooltip)
        self.annotation_pca = self.ax_pca.annotate(
            "", xy=(0,0), xytext=(15,15),
            textcoords="offset points",
            bbox=dict(boxstyle="round,pad=0.5", fc=self.BG_CARD, ec=self.BORDER, alpha=0.95),
            arrowprops=dict(arrowstyle="->", color=self.FG_MUTED)
        )
        self.annotation_pca.set_visible(False)
                    
        self.ax_pca.set_title(f"PCA - Agrupamiento de Pigmentos (Huella Química)\nPC1 ({var_pc1:.1f}%) vs PC2 ({var_pc2:.1f}%)", color=self.FG_MAIN, weight='bold')
        self.ax_pca.set_xlabel(f"Componente Principal 1 ({var_pc1:.1f}%)", color=self.FG_MUTED)
        self.ax_pca.set_ylabel(f"Componente Principal 2 ({var_pc2:.1f}%)", color=self.FG_MUTED)
        self.ax_pca.grid(True, linestyle="--", alpha=0.15, color=self.FG_MUTED)
        
        # Apply dark theme styling to all PCA plot elements
        self.ax_pca.set_facecolor(self.BG_CARD)
        self.ax_pca.tick_params(colors=self.FG_MUTED, which='both')
        for spine in self.ax_pca.spines.values():
            spine.set_color(self.BORDER)
            
        legend = self.ax_pca.legend(fontsize=9, loc='best')
        if legend:
            legend.get_frame().set_facecolor(self.BG_PANEL)
            legend.get_frame().set_edgecolor(self.BORDER)
            for text in legend.get_texts():
                text.set_color(self.FG_MAIN)
                
        self.canvas_pca.draw()

    def exportar_past_espectros(self):
        espectros_validos = {k: v for k, v in self.espectros_datos.items() if v.get('datos') is not None and not v.get('metadata', {}).get('es_suma', False)}
        if not espectros_validos:
            messagebox.showwarning("Advertencia", "No hay espectros válidos cargados para exportar.")
            return
            
        ruta_guardar = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
            title="Exportar Espectros para PAST"
        )
        if not ruta_guardar:
            return
            
        try:
            try:
                c_min = int(self.entry_canal_min.get())
            except Exception:
                c_min = 100
            try:
                c_max = int(self.entry_canal_max.get())
            except Exception:
                c_max = 1000
                
            rows = []
            primer_id = list(espectros_validos.keys())[0]
            df_primer = espectros_validos[primer_id]['datos']
            roi_energies = df_primer['Energia_keV'].values[c_min:c_max]
            channel_headers = [f"{e:.3f}keV" for e in roi_energies]
            
            for item_id, esp in espectros_validos.items():
                nombre = self.file_tree.item(item_id, "text").replace("📊 ", "")
                grupo = esp.get('grupo_padre', 'Sin clasificar').replace("Points ", "")
                counts = esp['datos']['Cuentas'].values[c_min:c_max]
                
                total_counts = counts.sum()
                if total_counts == 0:
                    total_counts = 1.0
                counts_normalized = counts / total_counts
                
                row_dict = {
                    'Muestra': nombre,
                    'Grupo': grupo
                }
                
                for i, col_name in enumerate(channel_headers):
                    if i < len(counts_normalized):
                        row_dict[col_name] = counts_normalized[i]
                    else:
                        row_dict[col_name] = 0.0
                        
                rows.append(row_dict)
                
            df_past = pd.DataFrame(rows)
            with pd.ExcelWriter(ruta_guardar, engine='openpyxl') as writer:
                df_past.to_excel(writer, sheet_name="PAST_Spectra", index=False)
                
            messagebox.showinfo("Éxito", f"Archivo exportado para PAST correctamente:\n{ruta_guardar}\n\nNota: Abre este archivo en PAST, marca 'Show/edit row labels' y 'Show/edit column labels'. Luego cambia la segunda columna ('Grupo') a tipo 'Group'.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar para PAST:\n{e}")

    def exportar_past_elementos(self):
        espectros_validos = {k: v for k, v in self.espectros_datos.items() if v.get('datos') is not None and not v.get('metadata', {}).get('es_suma', False)}
        if not espectros_validos:
            messagebox.showwarning("Advertencia", "No hay espectros válidos cargados para exportar.")
            return
            
        ruta_guardar = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
            title="Exportar Elementos para PAST"
        )
        if not ruta_guardar:
            return
            
        try:
            todos_elementos = set()
            muestras_data = []
            
            for item_id, esp in espectros_validos.items():
                nombre = self.file_tree.item(item_id, "text").replace("📊 ", "")
                grupo = esp.get('grupo_padre', 'Sin clasificar').replace("Points ", "")
                df = esp['datos']
                
                counts = df['Cuentas'].values
                energias = df['Energia_keV'].values
                fondo = lectura_espectros.calcular_fondo_snip(counts)
                picos = lectura_espectros.buscar_picos(energias, counts, fondo)
                
                elem_dict = {}
                for p in picos:
                    if p['elemento']:
                        sym = p['elemento'].split('(')[1].split(')')[0].strip() if '(' in p['elemento'] else str(p['elemento']).strip()
                        todos_elementos.add(sym)
                        elem_dict[sym] = elem_dict.get(sym, 0.0) + p['area_neta']
                        
                muestras_data.append({
                    'nombre': nombre,
                    'grupo': grupo,
                    'elementos': elem_dict
                })
                
            elementos_ordenados = sorted(list(todos_elementos))
            rows = []
            for m in muestras_data:
                row_dict = {
                    'Muestra': m['nombre'],
                    'Grupo': m['grupo']
                }
                for elem in elementos_ordenados:
                    row_dict[elem] = m['elementos'].get(elem, 0.0)
                rows.append(row_dict)
                
            df_past = pd.DataFrame(rows)
            with pd.ExcelWriter(ruta_guardar, engine='openpyxl') as writer:
                df_past.to_excel(writer, sheet_name="PAST_Elements", index=False)
                
            messagebox.showinfo("Éxito", f"Archivo de elementos exportado para PAST correctamente:\n{ruta_guardar}\n\nNota: Abre este archivo en PAST, marca 'Show/edit row labels' y 'Show/edit column labels'. Luego cambia la segunda columna ('Grupo') a tipo 'Group'.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar para PAST:\n{e}")

    def hover_pca(self, event):
        if not hasattr(self, 'sc_plots') or not hasattr(self, 'annotation_pca'):
            return
        vis = self.annotation_pca.get_visible()
        if event.inaxes == self.ax_pca:
            for sc, pts in self.sc_plots:
                cont, ind = sc.contains(event)
                if cont:
                    idx = ind['ind'][0]
                    name = pts['names'][idx]
                    category = sc.get_label()
                    x = pts['x'][idx]
                    y = pts['y'][idx]
                    
                    self.annotation_pca.xy = (x, y)
                    text = f"Muestra: {name}\nGrupo: {category}\nPC1: {x:.4f}\nPC2: {y:.4f}"
                    self.annotation_pca.set_text(text)
                    
                    self.annotation_pca.get_bbox_patch().set_facecolor(self.BG_CARD)
                    self.annotation_pca.get_bbox_patch().set_edgecolor(self.BORDER)
                    self.annotation_pca.set_color(self.FG_MAIN)
                    
                    self.annotation_pca.set_visible(True)
                    self.canvas_pca.draw_idle()
                    return
            if vis:
                self.annotation_pca.set_visible(False)
                self.canvas_pca.draw_idle()

    def on_pick_pca(self, event):
        if not hasattr(self, 'sc_plots'):
            return
        sc = event.artist
        for plot_sc, pts in self.sc_plots:
            if plot_sc == sc:
                idx = event.ind[0]
                item_id = pts['item_ids'][idx]
                
                # Seleccionar directamente el nodo de la muestra en el Treeview
                if self.file_tree.exists(item_id):
                    self.file_tree.selection_set(item_id)
                    self.file_tree.see(item_id)
                    self.on_tree_select(None)
                return

    def copiar_grafica_portapapeles(self):
        """
        Copia la figura actual del espectro (self.fig) al portapapeles del sistema operativo
        para que pueda pegarse directamente (Ctrl+V) en Word, PowerPoint, WhatsApp, etc., sin necesidad de guardar un archivo.
        """
        try:
            import io
            import os
            import time
            import subprocess
            from PIL import Image

            # Guardar el gráfico actual a un buffer de memoria en formato PNG de alta resolución
            buf = io.BytesIO()
            self.fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
            buf.seek(0)
            img = Image.open(buf)

            if os.name == 'nt':
                # Implementación nativa para Windows 32-bit y 64-bit usando Win32 API / ctypes (CF_DIB)
                import ctypes
                from ctypes import wintypes

                output = io.BytesIO()
                img.convert('RGB').save(output, 'BMP')
                data = output.getvalue()[14:]  # Omitir los 14 bytes del encabezado BMP para obtener formato CF_DIB

                user32 = ctypes.windll.user32
                kernel32 = ctypes.windll.kernel32

                # Configurar firma explícita de tipos (argtypes/restype) para evitar truncamiento de punteros a 32 bits en Python de 64 bits
                kernel32.GlobalAlloc.restype = wintypes.HGLOBAL
                kernel32.GlobalAlloc.argtypes = [wintypes.UINT, ctypes.c_size_t]
                kernel32.GlobalLock.restype = wintypes.LPVOID
                kernel32.GlobalLock.argtypes = [wintypes.HGLOBAL]
                kernel32.GlobalUnlock.argtypes = [wintypes.HGLOBAL]

                user32.OpenClipboard.argtypes = [wintypes.HWND]
                user32.OpenClipboard.restype = wintypes.BOOL
                user32.EmptyClipboard.restype = wintypes.BOOL
                user32.SetClipboardData.argtypes = [wintypes.UINT, wintypes.HANDLE]
                user32.SetClipboardData.restype = wintypes.HANDLE
                user32.CloseClipboard.restype = wintypes.BOOL

                CF_DIB = 8
                GMEM_MOVEABLE = 0x0002

                # Reintentar abrir el portapapeles por si otro programa lo tiene bloqueado momentáneamente
                opened = False
                for _ in range(5):
                    if user32.OpenClipboard(None):
                        opened = True
                        break
                    time.sleep(0.05)

                if not opened:
                    raise RuntimeError("El portapapeles está bloqueado por otra aplicación.")

                try:
                    user32.EmptyClipboard()
                    h_mem = kernel32.GlobalAlloc(GMEM_MOVEABLE, len(data))
                    if not h_mem:
                        raise MemoryError("No se pudo asignar memoria global para la imagen.")
                    
                    p_mem = kernel32.GlobalLock(h_mem)
                    if not p_mem:
                        raise MemoryError("No se pudo bloquear la memoria para escribir la imagen.")

                    ctypes.memmove(p_mem, data, len(data))
                    kernel32.GlobalUnlock(h_mem)
                    
                    if not user32.SetClipboardData(CF_DIB, h_mem):
                        raise RuntimeError("Falló al establecer los datos en el portapapeles.")
                finally:
                    user32.CloseClipboard()

                output.close()
                buf.close()
                messagebox.showinfo("Portapapeles", "¡Imagen de la gráfica copiada al portapapeles!\nPuedes pegarla directamente (Ctrl+V) en cualquier documento o programa.")

            else:
                # Implementación para Linux
                tmp_path = "/tmp/xrf_grafica_portapapeles.png"
                img.save(tmp_path)
                buf.close()

                copiado = False
                try:
                    subprocess.run(["xclip", "-selection", "clipboard", "-t", "image/png", "-i", tmp_path], check=True)
                    copiado = True
                except Exception:
                    try:
                        subprocess.run(["wl-copy", "<", tmp_path], shell=True, check=True)
                        copiado = True
                    except Exception:
                        pass

                if copiado:
                    messagebox.showinfo("Portapapeles", "¡Imagen de la gráfica copiada al portapapeles!")
                else:
                    messagebox.showinfo("Imagen Guardada", f"Se ha guardado la imagen en:\n{tmp_path}\n(Instala 'xclip' en Linux para copiado directo).")

        except Exception as e:
            messagebox.showerror("Error al Copiar", f"No se pudo copiar la imagen al portapapeles:\n{e}")

    def exportar_proyecto_rtx(self):
        """
        Exporta el estado actual del proyecto (todos los espectros cargados,
        sus grupos, canales, metadatos y elementos discernidos) a un archivo XML .rtx
        compatible con el software Bruker ARTAX.
        """
        if not self.espectros_datos:
            messagebox.showwarning("Sin Datos", "No hay espectros cargados para guardar en un proyecto .rtx.")
            return

        file_path = filedialog.asksaveasfilename(
            title="Guardar Proyecto Bruker ARTAX",
            defaultextension=".rtx",
            filetypes=[("Proyecto Bruker ARTAX (*.rtx)", "*.rtx"), ("Todos los archivos (*.*)", "*.*")]
        )
        
        if not file_path:
            return

        try:
            import xml.etree.ElementTree as ET
            from xml.dom import minidom

            root_elem = ET.Element("ClassInstance", attrib={"Type": "TRTProject", "Name": "RoentecProject"})
            grupos_nodes = {}

            # Recorrer todos los espectros en self.espectros_datos
            for item_id, esp in self.espectros_datos.items():
                datos_df = esp.get('datos')
                if datos_df is None:
                    continue
                
                meta = esp.get('metadata', {})
                
                # Excluir espectros de SUMA
                if meta.get('es_suma', False):
                    continue
                
                nombre_archivo = meta.get('nombre_archivo', 'Espectro')
                if 'espectro suma' in nombre_archivo.lower() or nombre_archivo.startswith('SUMA_'):
                    continue
                
                grupo_padre = esp.get('grupo_padre', '')
                
                # Crear contenedor de grupo en XML si aplica
                if grupo_padre and grupo_padre not in grupos_nodes:
                    base_node = ET.SubElement(root_elem, "ClassInstance", attrib={"Type": "TRTBase", "Name": str(grupo_padre)})
                    grupos_nodes[grupo_padre] = base_node
                
                parent_node = grupos_nodes[grupo_padre] if grupo_padre else root_elem
                nombre_clean = nombre_archivo.replace('.pdz', '').replace('.rtx', '')
                
                spec_node = ET.SubElement(parent_node, "ClassInstance", attrib={"Type": "TRTSpectrum", "Name": nombre_clean})
                
                # Metadatos del disparo
                live_time_ms = int(meta.get('live_time', 30.0) * 1000.0)
                voltage = meta.get('xray_voltage_kv', 40.0)
                current = meta.get('xray_filament_current', 100.0)
                num_channels = len(datos_df)
                ev_per_ch = meta.get('ev_per_channel', 20.0)
                calib_lin = ev_per_ch / 1000.0  # keV por canal
                
                header_node = ET.SubElement(spec_node, "Header")
                
                lt_node = ET.SubElement(header_node, "LifeTime")
                lt_node.text = str(live_time_ms)
                
                hv_node = ET.SubElement(header_node, "HighVoltage")
                hv_node.text = str(voltage)
                
                tc_node = ET.SubElement(header_node, "TubeCurrent")
                tc_node.text = str(current)
                
                cc_node = ET.SubElement(header_node, "ChannelCount")
                cc_node.text = str(num_channels)
                
                cl_node = ET.SubElement(header_node, "CalibLin")
                cl_node.text = f"{calib_lin:.6f}"
                
                sa_node = ET.SubElement(header_node, "SigmaAbs")
                sa_node.text = "0.0"

                # Guardar elementos discernidos si existen
                discernidos = esp.get('elementos_discernidos', [])
                if discernidos:
                    elem_node = ET.SubElement(header_node, "Elements")
                    elem_node.text = ",".join(discernidos)
                
                # Datos de cuentas del canal
                counts_array = datos_df['Cuentas'].values
                counts_str = ",".join(str(int(c)) for c in counts_array)
                
                ch_node = ET.SubElement(spec_node, "Channels")
                ch_node.text = counts_str

            # Formatear el XML estructurado
            xml_str = ET.tostring(root_elem, encoding="utf-8")
            parsed = minidom.parseString(xml_str)
            pretty_xml = parsed.toprettyxml(indent="  ", encoding="utf-8")

            with open(file_path, "wb") as f:
                f.write(pretty_xml)

            messagebox.showinfo("Éxito", f"Proyecto guardado correctamente en formato Bruker ARTAX:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Error de Exportación", f"No se pudo guardar el proyecto .rtx:\n{e}")

    def exportar_artax_filtrado(self):
        """
        Exporta un archivo Excel (.xls/.xlsx) con el formato exacto de Bruker ARTAX (todos.xls).
        Incluye las pestañas 'Parameter' y 'Points'.
        - 'Points': Contiene únicamente las muestras individuales reales (excluyendo cualquier fila de SUMA),
          y únicamente las columnas de los elementos seleccionados en el panel de Líneas de Referencia.
        - 'Parameter': Contiene los metadatos de medición y la lista de los elementos seleccionados.
        """
        if not self.espectros_datos:
            messagebox.showwarning("Advertencia", "No hay espectros cargados para exportar.")
            return

        elementos_seleccionados = [el for el, var in self.elementos_var.items() if var.get()]
        if not elementos_seleccionados:
            messagebox.showwarning(
                "Líneas no seleccionadas",
                "Por favor, selecciona al menos un elemento en el panel de 'Líneas de Referencia' para definir la línea base que deseas exportar."
            )
            return

        elementos_seleccionados.sort()

        # Filtrar únicamente espectros reales de muestra (sin sumas)
        espectros_reales = {}
        for item_id, esp in self.espectros_datos.items():
            if esp.get('datos') is None:
                continue
            meta = esp.get('metadata', {})
            is_suma = meta.get('es_suma', False) or item_id.endswith('/SUMA_AUTO') or esp.get('nombre_archivo', '').startswith('SUMA_')
            if not is_suma:
                espectros_reales[item_id] = esp

        if not espectros_reales:
            messagebox.showwarning("Advertencia", "No hay espectros individuales válidos para exportar.")
            return

        ruta_guardar = filedialog.asksaveasfilename(
            defaultextension=".xls",
            filetypes=[("Archivo Excel de Bruker ARTAX", "*.xls"), ("Todos los archivos", "*.*")],
            title="Guardar Espectros Filtrados en Formato ARTAX (todos.xls)",
            initialfile="todos_filtrado.xls"
        )
        if not ruta_guardar:
            return

        try:
            ARTAX_ELEMENT_LINES = {
                'Mg': [('Mg K12', 1.25)],
                'Al': [('Al K12', 1.49)],
                'Si': [('Si K12', 1.74)],
                'P':  [('P K12', 2.01)],
                'S':  [('S K12', 2.31)],
                'Cl': [('Cl K12', 2.62)],
                'Ar': [('Ar K12', 2.96)],
                'K':  [('K K12', 3.31)],
                'Ca': [('Ca K12', 3.69), ('Ca L1', 4.01)],
                'Ti': [('Ti K12', 4.51), ('Ti L1', 4.93)],
                'Cr': [('Cr K12', 5.41), ('Cr L1', 5.95)],
                'Mn': [('Mn K12', 5.90), ('Mn L1', 6.49)],
                'Fe': [('Fe K12', 6.40), ('Fe L1', 7.06)],
                'Co': [('Co K12', 6.93)],
                'Ni': [('Ni K12', 7.48), ('Ni L1', 8.26)],
                'Cu': [('Cu K12', 8.04), ('Cu L1', 8.90)],
                'Zn': [('Zn K12', 8.63), ('Zn L1', 9.57)],
                'Ga': [('Ga K12', 9.25)],
                'Ge': [('Ge K12', 9.88)],
                'As': [('As K12', 10.54)],
                'Se': [('Se K12', 11.22)],
                'Br': [('Br K12', 11.92)],
                'Rb': [('Rb K12', 13.39)],
                'Sr': [('Sr K12', 14.16), ('Sr L1', 1.81)],
                'Y':  [('Y K12', 14.96)],
                'Zr': [('Zr K12', 15.77)],
                'Nb': [('Nb K12', 16.61)],
                'Mo': [('Mo K12', 17.48)],
                'Ru': [('Ru K12', 19.28), ('Ru L1', 2.56)],
                'Rh': [('Rh K12', 20.21), ('Rh L1', 2.69)],
                'Pd': [('Pd K12', 21.18), ('Pd L1', 2.84)],
                'Ag': [('Ag K12', 22.16), ('Ag L1', 2.98)],
                'Cd': [('Cd K12', 23.17), ('Cd L1', 3.13)],
                'In': [('In K12', 24.21), ('In L1', 3.29)],
                'Sn': [('Sn K12', 25.27), ('Sn L1', 3.44)],
                'Sb': [('Sb K12', 26.36), ('Sb L1', 3.60)],
                'Ba': [('Ba L1', 4.47)],
                'Os': [('Os L1', 8.91), ('Os M1', 1.91)],
                'Au': [('Au L1', 9.71), ('Au M1', 2.12)],
                'Hg': [('Hg L1', 9.99)],
                'Pb': [('Pb L1', 10.55), ('Pb M1', 2.34)],
                'Bi': [('Bi L1', 10.84)],
                'Th': [('Th L1', 12.97), ('Th M1', 2.99)]
            }

            columnas_artax = []
            lineas_evaluar = []
            for el in elementos_seleccionados:
                lines = ARTAX_ELEMENT_LINES.get(el, [(f"{el} K12", 5.0)])
                for col_name, e_kev in lines:
                    columnas_artax.append(col_name)
                    lineas_evaluar.append((col_name, e_kev))

            wb = openpyxl.Workbook()
            ws_param = wb.active
            ws_param.title = "Parameter"

            primer_esp = next(iter(espectros_reales.values()))
            meta = primer_esp.get('metadata', {})

            high_voltage = meta.get('high_voltage', meta.get('xray_voltage_kv', 40))
            current = meta.get('current', meta.get('xray_filament_current', 11))
            live_time = meta.get('live_time', 24)
            filter_str = meta.get('filter', 'Ti/Al')

            param_rows = [
                ["Project:", None],
                [None, None],
                ["Ser.No.:", None],
                [None, None],
                ["Method:", "arqu encrym"],
                ["Measurement", None],
                ["High voltage/kV:", high_voltage],
                ["Current/µA:", current],
                ["Time/s:", live_time],
                ["Energy range/keV:", 0],
                ["Anode:", None],
                ["Filter:", filter_str],
                ["Optic:", "No optic"],
                ["Atmosphere:", "Air"],
                ["Evaluation", None],
                ["Corrections:", "Escape Backgr."],
                ["Stripping cycles:", 9],
                ["Elements:", " ".join(elementos_seleccionados) + " "],
                ["Deconvolution method:", "Bayes"],
            ]
            for r in param_rows:
                ws_param.append(r)

            for _ in range(14):
                ws_param.append([None, None])

            ws_param.append(["Values:", "Net area"])

            ws_points = wb.create_sheet(title="Points")
            headers = [None, None, None, "Muestra / Espectro"] + columnas_artax
            ws_points.append(headers)

            for item_id, esp in espectros_reales.items():
                nombre_archivo = esp.get('nombre_archivo', item_id)
                df_datos = esp.get('datos')

                # Limpiar nombre para dejar solo el identificador limpio (ej: ANALYZE_EMP-7060)
                nombre_limpio = os.path.basename(nombre_archivo)
                if '.' in nombre_limpio:
                    nombre_limpio = os.path.splitext(nombre_limpio)[0]
                if '@' in nombre_limpio:
                    nombre_limpio = nombre_limpio.split('@')[0]

                row_data = [None, None, None, nombre_limpio]

                if df_datos is not None and not df_datos.empty and 'Energia_keV' in df_datos and 'Cuentas' in df_datos:
                    energias = df_datos['Energia_keV'].values
                    counts = df_datos['Cuentas'].values
                    fondo = lectura_espectros.calcular_fondo_snip(counts)

                    for col_name, e_kev in lineas_evaluar:
                        idx = np.argmin(np.abs(energias - e_kev))
                        _, area_neta = lectura_espectros.calcular_area_neta_pico(counts, fondo, idx, window=7)
                        row_data.append(int(round(area_neta)))
                else:
                    for _ in lineas_evaluar:
                        row_data.append(0)

                ws_points.append(row_data)

            wb.save(ruta_guardar)

            messagebox.showinfo(
                "Exportación Exitosa",
                f"Se exportó correctamente el reporte estilo Bruker ARTAX:\n{ruta_guardar}\n\n"
                f"• Muestras procesadas: {len(espectros_reales)}\n"
                f"• Elementos filtrados: {len(elementos_seleccionados)} ({', '.join(elementos_seleccionados)})\n"
                f"• Columnas de líneas: {len(columnas_artax)}"
            )

        except Exception as e:
            messagebox.showerror("Error al Exportar ARTAX", f"Ocurrió un error inesperado al guardar el archivo:\n{e}")

    def exportar_datos(self):
        if not self.espectros_datos:
            messagebox.showwarning("Advertencia", "No hay espectros cargados para exportar.")
            return
            
        ruta_guardar = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
            title="Guardar espectros en Excel Universal"
        )
        if not ruta_guardar:
            return
            
        try:
            # Crear DataFrame consolidado para Datos Brutos (Canales de Energía vs Cuentas)
            primer_id = list(self.espectros_datos.keys())[0]
            df_base = self.espectros_datos[primer_id]['datos'][['Energia_keV']].copy()
            
            for item_id, esp in self.espectros_datos.items():
                nombre_col = self.file_tree.item(item_id, "text").replace("📊 ", "")
                df_base[nombre_col] = esp['datos']['Cuentas']
            
            summary_rows = []
            peaks_rows = []
            
            for item_id, esp in self.espectros_datos.items():
                nombre_clean = self.file_tree.item(item_id, "text").replace("📊 ", "")
                df = esp['datos']
                meta = esp['metadata']
                
                counts = df['Cuentas'].values
                energias = df['Energia_keV'].values
                
                # Calcular fondo SNIP e identificar picos
                fondo = lectura_espectros.calcular_fondo_snip(counts)
                picos = lectura_espectros.buscar_picos(energias, counts, fondo)
                
                # Recopilar elementos detectados automáticamente y llenar tabla de picos
                elementos_detectados_auto = []
                for p in picos:
                    elem_display = p['elemento'] if p['elemento'] else 'Desconocido'
                    if p['elemento']:
                        sym = p['elemento'].split('(')[1].split(')')[0].strip() if '(' in p['elemento'] else str(p['elemento']).strip()
                        elementos_detectados_auto.append(sym)
                    
                    peaks_rows.append({
                        'Espectro': nombre_clean,
                        'Elemento': elem_display,
                        'Energía (keV)': p['energia_kev'],
                        'Cuentas Brutas': p['cuentas_brutas'],
                        'Cuentas Netas': p['cuentas_netas'],
                        'Área Neta': p['area_neta'],
                        'Área Relativa %': p['area_relativa']
                    })
                
                elementos_detectados_auto = list(set(elementos_detectados_auto))
                
                # Obtener elementos discernidos por el usuario (o default auto)
                discernidos = esp.get('elementos_discernidos')
                if discernidos is None:
                    discernidos = elementos_detectados_auto
                    esp['elementos_discernidos'] = discernidos
                    
                # Sugerencias de pigmentos basadas en elementos discernidos
                sugerencias = lectura_espectros.sugerir_pigmentos(discernidos)
                sugerencias_str_list = [f"{s['pigmento']} ({s['color']}): {s['justificacion']}" for s in sugerencias]
                
                summary_rows.append({
                    'ID Espectro': nombre_clean,
                    'Grupo / Categoría': esp.get('grupo_padre', 'Sin clasificar'),
                    'Tiempo Vivo (s)': meta.get('live_time', 0.0),
                    'Voltaje Tubo (kV)': meta.get('xray_voltage_kv', 0.0),
                    'Corriente Tubo (uA)': meta.get('xray_filament_current', 0.0),
                    'Elementos Detectados (Auto)': ", ".join(elementos_detectados_auto),
                    'Elementos Discernidos (Humano)': ", ".join(discernidos),
                    'Sugerencias de Pigmentos (Discernidos)': " | ".join(sugerencias_str_list)
                })
                
            df_summary = pd.DataFrame(summary_rows)
            df_peaks = pd.DataFrame(peaks_rows)
            
            with pd.ExcelWriter(ruta_guardar, engine='openpyxl') as writer:
                # Escribir cada DataFrame a su pestaña correspondiente
                df_base.to_excel(writer, sheet_name="Datos Brutos", index=False)
                df_summary.to_excel(writer, sheet_name="Resumen de Muestras", index=False)
                df_peaks.to_excel(writer, sheet_name="Detalle de Picos", index=False)
                
            messagebox.showinfo("Éxito", f"Datos consolidados exportados correctamente en:\n{ruta_guardar}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron exportar los datos:\n{e}")

    def sanitizar_nombre_hoja(self, nombre):
        # Nombres de hojas en Excel tienen un máximo de 31 caracteres
        # y no pueden contener caracteres especiales como: \ / ? * : [ o ]
        caracteres_prohibidos = ['\\', '/', '?', '*', ':', '[', ']']
        for c in caracteres_prohibidos:
            nombre = nombre.replace(c, '_')
        return nombre[:30] # Limitamos a 30 para seguridad

    def obtener_nombre_hoja_unico(self, nombre, nombres_existentes):
        sanitizado = self.sanitizar_nombre_hoja(nombre)
        base = sanitizado
        contador = 1
        while sanitizado.lower() in nombres_existentes:
            sufijo = f"_{contador}"
            limite = 31 - len(sufijo)
            sanitizado = base[:limite] + sufijo
            contador += 1
        nombres_existentes.add(sanitizado.lower())
        return sanitizado

    def exportar_datos_con_graficas(self):
        if not self.espectros_datos:
            messagebox.showwarning("Advertencia", "No hay espectros cargados para exportar.")
            return
            
        ruta_guardar = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
            title="Guardar reporte con gráficas por hoja"
        )
        if not ruta_guardar:
            return
            
        # Cambiar el cursor a reloj/espera mientras se procesa
        self.root.config(cursor="watch")
        self.root.update()
        
        try:
            import io
            import re
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
            from openpyxl.drawing.image import Image
            
            wb = openpyxl.Workbook()
            # Eliminar la primera hoja por defecto para agregar las nuestras
            if len(wb.sheetnames) > 0:
                wb.remove(wb.active)
                
            nombres_existentes = set()
            
            # Estilos de celda premium
            header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid') # Azul oscuro institucional
            section_font = Font(name='Segoe UI', size=12, bold=True, color='1F4E79')
            bold_font = Font(name='Segoe UI', size=10, bold=True)
            regular_font = Font(name='Segoe UI', size=10)
            
            # Separar espectros reales y sumas
            espectros_reales = []
            espectros_sumas = []
            
            for item_id, esp in self.espectros_datos.items():
                is_suma = esp.get('metadata', {}).get('es_suma', False)
                if is_suma:
                    espectros_sumas.append((item_id, esp))
                else:
                    espectros_reales.append((item_id, esp))
                    
            # Ordenar espectros reales por número ascendente en su nombre
            def obtener_numero_nombre(item):
                item_id, _ = item
                nombre = self.file_tree.item(item_id, "text").replace("📊 ", "")
                match = re.search(r'\d+', nombre)
                return int(match.group()) if match else 0
                
            espectros_reales.sort(key=obtener_numero_nombre)
            
            # Ordenar sumas alfabéticamente por su nombre
            def obtener_nombre_limpio(item):
                item_id, _ = item
                return self.file_tree.item(item_id, "text").replace("📊 ", "")
                
            espectros_sumas.sort(key=obtener_nombre_limpio)
            
            # Unir las listas, reales primero y sumas al final
            espectros_ordenados = espectros_reales + espectros_sumas
            
            # Recorrer cada espectro cargado en orden
            for item_id, esp in espectros_ordenados:
                nombre_clean = self.file_tree.item(item_id, "text").replace("📊 ", "")
                df = esp['datos']
                meta = esp['metadata']
                
                counts = df['Cuentas'].values
                energias = df['Energia_keV'].values
                
                # Calcular fondo SNIP e identificar picos
                fondo = lectura_espectros.calcular_fondo_snip(counts)
                picos = lectura_espectros.buscar_picos(energias, counts, fondo)
                
                # Crear la pestaña para esta muestra
                sheet_name = self.obtener_nombre_hoja_unico(nombre_clean, nombres_existentes)
                ws = wb.create_sheet(title=sheet_name)
                ws.sheet_view.showGridLines = True
                
                # Escribir encabezados
                cell_a1 = ws.cell(row=1, column=1, value="Energía (keV)")
                cell_a1.font = header_font
                cell_a1.fill = header_fill
                
                cell_b1 = ws.cell(row=1, column=2, value="Cuentas")
                cell_b1.font = header_font
                cell_b1.fill = header_fill
                
                cell_c1 = ws.cell(row=1, column=3, value="Elemento")
                cell_c1.font = header_font
                cell_c1.fill = header_fill
                
                # Mapear los picos identificados a los canales de energía correspondientes
                picos_map = {}
                for p in picos:
                    if p['elemento']:
                        idx_cercano = np.abs(energias - p['energia_kev']).argmin()
                        picos_map[idx_cercano] = p['elemento']
                
                # Escribir los valores
                for r_idx, (eng, cnt) in enumerate(zip(energias, counts)):
                    r = 2 + r_idx
                    ws.cell(row=r, column=1, value=eng).font = regular_font
                    ws.cell(row=r, column=2, value=cnt).font = regular_font
                    
                    elem_val = picos_map.get(r_idx, "")
                    ws.cell(row=r, column=3, value=elem_val).font = regular_font
                
                # Definir anchos de columna fijos y compactos
                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 12
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 3
                
                # -------------------------------------------------------------
                # GENERAR GRÁFICO ASOCIADO (MATPLOTLIB)
                # -------------------------------------------------------------
                # Usar estilo limpio y claro para impresión en Excel
                fig_tmp = Figure(figsize=(8, 4.5), dpi=100)
                ax_tmp = fig_tmp.add_subplot(111)
                
                # Graficar espectro y fondo
                ax_tmp.plot(energias, counts, color='#1f77b4', label='Espectro Bruto', linewidth=1.2)
                ax_tmp.plot(energias, fondo, color='#ff7f0e', linestyle='--', label='Fondo (SNIP)', linewidth=1.2)
                
                # Graficar picos
                ax_tmp.plot([p['energia_kev'] for p in picos], [p['cuentas_brutas'] for p in picos], 'ro', markersize=4)
                for p in picos:
                    if p['elemento'] and self.mostrar_etiquetas_picos_var.get():
                        ax_tmp.text(
                            p['energia_kev'], 
                            p['cuentas_brutas'] + (max(counts) * 0.02), 
                            p['elemento'], 
                            fontsize=8, 
                            rotation=90, 
                            verticalalignment='bottom', 
                            horizontalalignment='center',
                            color='#d62728',
                            weight='bold'
                        )
                
                # Graficar líneas de referencia si están activas
                rango_max = self.obtener_rango_max()
                    
                for el, var in self.elementos_var.items():
                    if var.get():
                        for key, val in lectura_espectros.XRF_ELEMENTS.items():
                            if key.startswith(f"{el}_") or key == el:
                                energy, line, label = val
                                if energy <= rango_max:
                                    ax_tmp.axvline(x=energy, color='gray', linestyle=':', linewidth=0.8)
                                    ax_tmp.text(
                                        energy, 
                                        ax_tmp.get_ylim()[1] * 0.95, 
                                        f"{el} {line}", 
                                        color='gray', 
                                        fontsize=7, 
                                        rotation=90, 
                                        verticalalignment='top', 
                                        horizontalalignment='right'
                                    )
                
                ax_tmp.set_title(f"Espectro: {meta.get('nombre_archivo', nombre_clean)}")
                ax_tmp.set_xlabel("Energía (keV)")
                ax_tmp.set_ylabel("Cuentas")
                ax_tmp.set_xlim(0, rango_max)
                ax_tmp.grid(True, linestyle="--", alpha=0.3)
                ax_tmp.legend(fontsize=9, loc='upper right')
                
                # Guardar el gráfico en un stream de memoria BytesIO
                img_buf = io.BytesIO()
                fig_tmp.savefig(img_buf, format='png', bbox_inches='tight')
                img_buf.seek(0)
                
                # Cargar imagen en openpyxl e insertarla en la celda E2 (dejando D como espaciado)
                img = Image(img_buf)
                ws.add_image(img, 'E2')
                
                # Si es una hoja de suma, agregar la tabla de archivos utilizados en la columna E debajo del gráfico
                is_suma = esp.get('metadata', {}).get('es_suma', False)
                if is_suma:
                    cell_title = ws.cell(row=27, column=5, value="Archivos Utilizados") # Columna E, Fila 27
                    cell_title.font = header_font
                    cell_title.fill = header_fill
                    
                    folder_id = os.path.dirname(item_id)
                    children = self.file_tree.get_children(folder_id)
                    archivos_utilizados = []
                    for child in children:
                        if child in self.espectros_datos:
                            child_esp = self.espectros_datos[child]
                            if not child_esp.get('metadata', {}).get('es_suma', False):
                                name = self.file_tree.item(child, "text").replace("📊 ", "")
                                archivos_utilizados.append(name)
                                
                    for idx, name in enumerate(archivos_utilizados):
                        cell_file = ws.cell(row=28 + idx, column=5, value=name)
                        cell_file.font = regular_font
                        
                    ws.column_dimensions['E'].width = 35
                
            wb.save(ruta_guardar)
            messagebox.showinfo("Éxito", f"Reporte completo con gráficas exportado correctamente en:\n{ruta_guardar}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron exportar los datos con gráficos:\n{e}")
        finally:
            # Restaurar el cursor normal
            self.root.config(cursor="")
            self.root.update()

    def popular_guia_teorica(self):
        contenido = """================================================================================
GUÍA DE REFERENCIA DE PIGMENTOS HISTÓRICOS Y ARQUEOLÓGICOS (TEORÍA XRF)
================================================================================

1. TABLA DE CORRELACIONES ELEMENTALES
--------------------------------------------------------------------------------
Color       | Pigmento Histórico / Compuesto     | Elementos Clave (XRF)
------------+------------------------------------+------------------------------
ROJO        | Bermellón / Cinabrio               | Hg (Mercurio)
            | Minio / Plomo Rojo                 | Pb (Plomo)
            | Realgar / Pararealgar              | As (Arsénico) + S (Azufre)
            | Hematita / Ocre Rojo               | Fe (Hierro)
------------+------------------------------------+------------------------------
AMARILLO    | Oropimente                         | As (Arsénico) + S (Azufre)
            | Amarillo de Plomo-Estaño           | Pb (Plomo) + Sn (Estaño)
            | Goethita / Ocre Amarillo           | Fe (Hierro)
------------+------------------------------------+------------------------------
AZUL        | Azurita                            | Cu (Cobre)
            | Esmalte (Smalt)                    | Co (Cobalto) + Si + K
            | Azul Egipcio                       | Cu (Cobre) + Si + Ca
            | Azul de Prusia (Moderno)           | Fe (Hierro)
------------+------------------------------------+------------------------------
VERDE       | Malaquita / Cardenillo             | Cu (Cobre)
            | Verde Esmeralda (París)            | Cu (Cobre) + As (Arsénico)
            | Tierra Verde (Terra Verte)         | Fe (Hierro) + Si + Al + K
------------+------------------------------------+------------------------------
BLANCO      | Blanco de Plomo                    | Pb (Plomo)
            | Yeso                               | Ca (Calcio) + S (Azufre)
            | Calcita / Tiza                     | Ca (Calcio)
            | Blanco de Titanio (Moderno)        | Ti (Titanio) (Sin Fe)
            | Blanco de Zinc (Moderno)           | Zn (Zinc)
------------+------------------------------------+------------------------------
NEGRO       | Negro de Manganeso                 | Mn (Manganeso)
            | Negro de Hierro / Magnetita        | Fe (Hierro) [Alto]
            | Negro de Hueso / Marfil            | Ca (Calcio) + P (Fósforo)
            | Negro de Carbón / Carbón Vegetal   | Orgánico (XRF no detecta C,
            |                                    | se nota por alta dispersión)
------------+------------------------------------+------------------------------
CAFÉ/MARRÓN | Tierra de Sombra (Umber)           | Fe (Hierro) + Mn (Manganeso)
            | Tierra Marrón / Goethita           | Fe (Hierro)
--------------------------------------------------------------------------------

2. ADVERTENCIAS CRÍTICAS E INSTRUMENTALES (EL HAZ DEL XRF)
--------------------------------------------------------------------------------
* ARTEFACTOS DEL HAZ DEL EQUIPO (RODIO Y PALADIO):
  Los picos de Rodio (Rh) y Paladio (Pd) detectados en los espectros NO son parte
  del objeto analizado. Se deben a la dispersión (elástica e inelástica) del haz 
  del propio tubo de rayos X del equipo. Deben ser IGNORADOS como pigmentos.

* SOLAPAMIENTO PLOMO-ARSÉNICO (Pb-As):
  El pico principal del Arsénico (As Ka a 10.54 keV) coincide casi de forma exacta 
  con el del Plomo (Pb La a 10.55 keV). Si se usa Blanco de Plomo, puede ocultar 
  la presencia de Oropimente/Realgar (As).
  --> Sugerencia: Verificar la presencia del pico secundario As Kb a 11.73 keV.

* SOLAPAMIENTO AZUFRE-PLOMO (S-Pb):
  Las líneas del Azufre (S Ka a 2.31 keV) se solapan fuertemente con la línea Pb M 
  (2.34 keV) si hay plomo. Esto dificulta la confirmación directa de Azufre en 
  yeso o sulfuros si el blanco de plomo está presente.

3. SUGERENCIAS PARA LA INTERPRETACIÓN CIENTÍFICA EN REPORTES
--------------------------------------------------------------------------------
A) EFECTOS DE PENETRACIÓN (MATRIZ / SOPORTE):
   El haz del XRF penetra varias décimas de milímetro (e incluso milímetros en 
   materiales de bajo número atómico). Por ello, elementos como el Calcio (Ca) 
   o Azufre (S) detectados en zonas pintadas a menudo provienen de la capa de 
   preparación (yeso/calcita) o del propio soporte subyacente. 
   --> Práctica recomendada: Medir siempre un punto "blanco" o "base" del soporte 
       para sustraerlo mentalmente de las lecturas.

B) COMPLEMENTARIEDAD MOLECULAR:
   Dado que el XRF solo identifica átomos (elementos) y no enlaces moleculares, 
   la presencia de Cobre (Cu) en un área verde es compatible tanto con Malaquita 
   como con Cardenillo o resinato de cobre. Se recomienda cruzar los datos de XRF 
   con espectroscopía Raman o FTIR para la identificación de la fase mineral 
   o compuestos orgánicos.

C) FECHADO POR ANACRONISMOS:
   La detección de elementos como el Titanio (Ti) (Blanco de Titanio) o el Zinc 
   (Zn) (Blanco de Zinc) en obras que se suponen anteriores al siglo XIX es un 
   indicador muy fuerte de restauraciones modernas o potenciales falsificaciones.
================================================================================"""
        self.guia_text.config(state="normal")
        self.guia_text.delete("1.0", tk.END)
        self.guia_text.insert(tk.END, contenido)
        self.guia_text.config(state="disabled")

# Helper para Tooltips
class Tooltip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip_window = None
        self.widget.bind("<Enter>", self.show_tip)
        self.widget.bind("<Leave>", self.hide_tip)

    def show_tip(self, event=None):
        if self.tip_window or not self.text:
            return
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 5
        self.tip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        
        label = tk.Label(
            tw, 
            text=self.text, 
            justify="left", 
            background="#1E1E1E", 
            foreground="#D4D4D4", 
            relief="flat", 
            font=("Helvetica", 9),
            padx=8,
            pady=4,
            highlightthickness=1,
            highlightbackground="#3E3E42"
        )
        label.pack()

    def hide_tip(self, event=None):
        tw = self.tip_window
        self.tip_window = None
        if tw:
            tw.destroy()

# Bloque de ejecución principal
if __name__ == "__main__":
    root = TkinterDnD.Tk()
    app = XRFProcessorGUI(root)
    root.mainloop()
